"""
Budget Template Manager

Handles the creation, population, and management of Oracle Budget FBDI templates.
Supports dynamic segment mapping (SEGMENT1-SEGMENT30) for flexible chart of accounts.
"""

import os
import time
import shutil
from pathlib import Path
from typing import Dict, List, Any, Optional
import pandas as pd
from openpyxl import load_workbook

from oracle_fbdi_integration.core.file_utils import excel_to_csv_and_zip
from account_and_entitys.oracle import OracleSegmentMapper


class BudgetTemplateManager:
    """Manages Oracle Budget FBDI template operations."""

    def __init__(self, template_path: str):
        """
        Initialize the Budget Template Manager.

        Args:
            template_path: Path to the BudgetImportTemplate.xlsm file
        """
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template file not found: {template_path}")
        
        self.mapper = OracleSegmentMapper()

    def create_clean_template(self, output_path: Optional[str] = None) -> str:
        """
        Create a clean copy of the template with empty XCC_BUDGET_INTERFACE data rows.

        Args:
            output_path: Path for the clean template (auto-generated if not provided)

        Returns:
            Path to the clean template file
        """
        if output_path is None:
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            output_path = self.template_path.parent / f"BudgetImportTemplate_Clean_{timestamp}.xlsm"
        else:
            output_path = Path(output_path)

        # Create copy and clear data
        shutil.copy2(self.template_path, output_path)
        print(f"Created clean template copy: {output_path}")

        wb = load_workbook(output_path)
        if "XCC_BUDGET_INTERFACE" in wb.sheetnames:
            budget_sheet = wb["XCC_BUDGET_INTERFACE"]
            # Keep rows 1-4 (instructions and headers), delete data rows from row 5 onwards
            if budget_sheet.max_row > 4:
                budget_sheet.delete_rows(5, budget_sheet.max_row - 4)
                print("Cleared data rows from XCC_BUDGET_INTERFACE sheet")
            wb.save(output_path)
            wb.close()
        else:
            wb.close()
            raise ValueError("XCC_BUDGET_INTERFACE sheet not found in template")

        return str(output_path)

    def fill_template(
        self,
        template_path: str,
        budget_data: List[Dict[str, Any]],
        output_path: Optional[str] = None,
        auto_zip: bool = False,
    ) -> str:
        """
        Fill a template with budget entry data.

        Args:
            template_path: Path to the clean template
            budget_data: List of dictionaries containing budget entry data
            output_path: Path for the filled template (optional)
            auto_zip: Whether to automatically create ZIP file

        Returns:
            Path to the filled template file (or ZIP file if auto_zip=True)
        """
        template_file = Path(template_path)
        if not template_file.exists():
            raise FileNotFoundError(f"Template file not found: {template_path}")
        if not budget_data:
            raise ValueError("No budget data provided")

        if output_path is None:
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            output_path = template_file.parent / f"XccBudgetInterface_Filled_{timestamp}.xlsm"
        else:
            output_path = Path(output_path)

        # Create copy and fill with data
        shutil.copy2(template_file, output_path)

        wb = load_workbook(output_path)
        budget_sheet = wb["XCC_BUDGET_INTERFACE"]

        # Get headers from row 4
        headers = []
        for col_num in range(1, budget_sheet.max_column + 1):
            header_cell = budget_sheet.cell(row=4, column=col_num)
            if header_cell.value is not None and str(header_cell.value).strip():
                clean_header = str(header_cell.value).lstrip('*').strip()
                headers.append(clean_header)
            else:
                break

        print(f"Found {len(headers)} headers in template")
        print(f"Filling template with {len(budget_data)} budget entries")

        # Fill data starting from row 5
        for row_idx, entry in enumerate(budget_data, start=5):
            for col_idx, header in enumerate(headers, start=1):
                value = entry.get(header, "")
                cell = budget_sheet.cell(row=row_idx, column=col_idx)
                if value is not None:
                    cell.value = value
                else:
                    cell.value = ""

        wb.save(output_path)
        wb.close()

        print(f"Template filled with data: {output_path}")

        # Optionally create ZIP file
        if auto_zip:
            zip_path = str(output_path).replace('.xlsm', '.zip')
            try:
                zip_result = excel_to_csv_and_zip(str(output_path), zip_path)
                print(f"ZIP file created: {zip_result}")
                return zip_result
            except Exception as e:
                print(f"Warning: Failed to create ZIP file: {e}")
                return str(output_path)

        return str(output_path)

    def create_from_scratch(
        self,
        budget_data: List[Dict[str, Any]],
        output_name: Optional[str] = None,
        auto_zip: bool = True,
    ) -> str:
        """
        Complete workflow: Create clean template, fill with data, and optionally ZIP.

        Args:
            budget_data: List of dictionaries containing budget entry data
            output_name: Base name for output files (optional)
            auto_zip: Whether to create ZIP file automatically

        Returns:
            Path to the final file (Excel or ZIP)
        """
        timestamp = time.strftime("%Y%m%d_%H%M%S")

        if output_name is None:
            output_name = f"XccBudgetInterface_{timestamp}"

        try:
            # Step 1: Create clean template
            clean_template = self.create_clean_template()

            # Step 2: Fill with data
            filled_template_path = self.template_path.parent / f"{output_name}.xlsm"
            result_path = self.fill_template(
                clean_template,
                budget_data,
                str(filled_template_path),
                auto_zip=auto_zip,
            )

            # Clean up temporary clean template
            Path(clean_template).unlink(missing_ok=True)

            return result_path

        except Exception as e:
            print(f"Error in budget creation workflow: {e}")
            raise


def create_budget_entry_data(
    transfers,
    transaction_id: int,
    source_budget_type: str = "HYPERION",
    source_budget_name: str = "MIC_HQ_MONTHLY",
    period_name: str = "1-25",
    currency_code: str = "AED",
) -> List[Dict[str, Any]]:
    """
    Create budget entry data with dynamic segment support from transaction transfers.

    Args:
        transfers: List of XX_TransactionTransfer objects with segment relationships
        transaction_id: Transaction identifier for budget entry naming
        source_budget_type: Budget type (e.g., "HYPERION", "ORACLE")
        source_budget_name: Source budget identifier
        period_name: Period name (e.g., "Sep-25", "Oct-25")
        currency_code: Currency code (default: AED)

    Returns:
        List of budget entries with dynamic SEGMENT1-SEGMENT30 columns

    Notes:
        - Uses OracleSegmentMapper to dynamically build segment columns
        - Creates separate entries for FROM (negative) and TO (positive) amounts
        - Supports 2-30 segments based on transaction configuration
    """
    mapper = OracleSegmentMapper()


    SOURCE_BUDGET_TYPE=os.getenv("ORACLE_SOURCE_BUDGET_TYPE")
    currency_code=os.getenv("ORACLE_CURRENCY_CODE")


    # Generate batch identifiers
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    budget_name = f"{source_budget_name}_{transaction_id}_{timestamp}"

    budget_entries = []
    line_number = 1
    ORACLE_BUDUGET_NAME=os.getenv("ORACLE_BUDUGET_NAME")
    for Budget_catgoray in ORACLE_BUDUGET_NAME.split(","):
        source_budget_name = Budget_catgoray
        budget_name = f"{source_budget_name}_{transaction_id}"
        for transfer in transfers:
            from_amount = getattr(transfer, 'from_center', 0) or 0
            if from_amount > 0:
                # Base budget entry
                budget_entry = {
                    "Source Budget Type": SOURCE_BUDGET_TYPE,
                    "Source Budget Name": source_budget_name,
                    "Budget Entry Name": budget_name,
                    "Line Number": line_number,
                    "Amount": -1 * from_amount,  # Negative for FROM side
                    "Currency Code": currency_code,
                    "Period Name": period_name,
                }

                # Add dynamic segment columns using mapper
                budget_entry_with_segments = mapper.build_fbdi_row(
                    transaction_transfer=transfer,
                    base_row=budget_entry,
                    include_from_to='from',
                    fill_all=False # Use FROM segments
                )

                budget_entries.append(budget_entry_with_segments)
                line_number += 1

            # Process TO side (positive amount - credit)
            to_amount = getattr(transfer, 'to_center', 0) or 0
            if to_amount > 0:
                # Base budget entry
                budget_entry = {
                    "Source Budget Type": SOURCE_BUDGET_TYPE,
                    "Source Budget Name": source_budget_name,
                    "Budget Entry Name": budget_name,
                    "Line Number": line_number,
                    "Amount": to_amount,  # Positive for TO side
                    "Currency Code": currency_code,
                    "Period Name": period_name,
                }

                # Add dynamic segment columns using mapper
                budget_entry_with_segments = mapper.build_fbdi_row(
                    transaction_transfer=transfer,
                    base_row=budget_entry,
                    include_from_to='to'  # Use TO segments
                )

                budget_entries.append(budget_entry_with_segments)
                line_number += 1

    return budget_entries


def create_budget_entry_with_segments(
    segment_values: Dict[int, str],
    amount: float,
    budget_name: str,
    line_number: int = 1,
    source_budget_type: str = "HYPERION",
    source_budget_name: str = "MIC_HQ_MONTHLY",
    period_name: str = "Sep-25",
    currency_code: str = "AED",
) -> Dict[str, Any]:
    """
    Create a single budget entry with custom segment values.

    Args:
        segment_values: Dict mapping segment_type_id to segment code
                       Example: {1: 'E001', 2: 'A100', 3: 'P001'}
        amount: Budget amount (positive for increase, negative for decrease)
        budget_name: Budget entry name identifier
        line_number: Line number within budget entry
        source_budget_type: Budget type (e.g., "HYPERION", "ORACLE")
        source_budget_name: Source budget identifier
        period_name: Period name (e.g., "Sep-25", "Oct-25")
        currency_code: Currency code (default: AED)

    Returns:
        Dictionary with all FBDI columns including dynamic SEGMENT1-SEGMENT30

    Example:
        >>> entry = create_budget_entry_with_segments(
        ...     segment_values={1: 'E001', 2: 'A100', 3: 'P001'},
        ...     amount=10000.00,
        ...     budget_name="Q3_BUDGET_TRANSFER",
        ...     line_number=1
        ... )
    """
    mapper = OracleSegmentMapper()

    # Base budget entry structure
    base_entry = {
        "Source Budget Type": source_budget_type,
        "Source Budget Name": source_budget_name,
        "Budget Entry Name": budget_name,
        "Line Number": line_number,
        "Amount": amount,
        "Currency Code": currency_code,
        "Period Name": period_name,
    }

    # Build segment columns dynamically
    for segment_type_id, segment_code in segment_values.items():
        oracle_field_name = mapper.get_oracle_field_name(segment_type_id)
        if oracle_field_name:
            base_entry[oracle_field_name] = str(segment_code)

    # Fill remaining segments with empty values for unconfigured segments
    all_segments = mapper.build_all_oracle_fields(num_segments=30)

    for segment_field in all_segments:
        if segment_field not in base_entry:
            base_entry[segment_field] = ""

    return base_entry


def create_budget_transfer_pair(
    from_segments: Dict[int, str],
    to_segments: Dict[int, str],
    amount: float,
    budget_name: str,
    period_name: str = "Sep-25",
    starting_line_number: int = 1,
    source_budget_type: str = "HYPERION",
    source_budget_name: str = "MIC_HQ_MONTHLY",
    currency_code: str = "AED",
) -> List[Dict[str, Any]]:
    """
    Create a balanced pair of budget entries (FROM negative + TO positive).

    Args:
        from_segments: Segment values for FROM entry (debit/decrease)
        to_segments: Segment values for TO entry (credit/increase)
        amount: Transfer amount (positive value)
        budget_name: Budget entry identifier
        period_name: Period name (e.g., "Sep-25")
        starting_line_number: Starting line number for entries
        source_budget_type: Budget type (e.g., "HYPERION", "ORACLE")
        source_budget_name: Source budget identifier
        currency_code: Currency code (default: AED)

    Returns:
        List of 2 budget entries (FROM negative, TO positive)

    Example:
        >>> entries = create_budget_transfer_pair(
        ...     from_segments={1: 'E001', 2: 'A100'},
        ...     to_segments={1: 'E002', 2: 'A200'},
        ...     amount=5000.00,
        ...     budget_name="MONTHLY_REALLOCATION"
        ... )
    """
    # Create FROM entry (negative amount)
    from_entry = create_budget_entry_with_segments(
        segment_values=from_segments,
        amount=-1 * amount,  # Negative for decrease
        budget_name=budget_name,
        line_number=starting_line_number,
        period_name=period_name,
        source_budget_type=source_budget_type,
        source_budget_name=source_budget_name,
        currency_code=currency_code,
    )

    # Create TO entry (positive amount)
    to_entry = create_budget_entry_with_segments(
        segment_values=to_segments,
        amount=amount,  # Positive for increase
        budget_name=budget_name,
        line_number=starting_line_number + 1,
        period_name=period_name,
        source_budget_type=source_budget_type,
        source_budget_name=source_budget_name,
        currency_code=currency_code,
    )

    return [from_entry, to_entry]
