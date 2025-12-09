"""
Journal Template Manager

Handles the creation, population, and management of Oracle GL Journal FBDI templates.
Supports dynamic segment mapping (SEGMENT1-SEGMENT30) for flexible chart of accounts.
"""

import os
import time
import shutil
from pathlib import Path
from typing import Dict, List, Any, Optional
import pandas as pd
from openpyxl import load_workbook

from budget_management.models import xx_BudgetTransfer
from oracle_fbdi_integration.core.file_utils import excel_to_csv_and_zip
from account_and_entitys.oracle import OracleSegmentMapper
from transaction.models import xx_TransactionTransfer


class JournalTemplateManager:
    """Manages Oracle GL Journal FBDI template operations."""

    def __init__(self, template_path: str):
        """
        Initialize the Journal Template Manager.

        Args:
            template_path: Path to the JournalImportTemplate.xlsm file
        """
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template file not found: {template_path}")
        
        self.mapper = OracleSegmentMapper()

    def create_clean_template(self, output_path: Optional[str] = None) -> str:
        """
        Create a clean copy of the template with empty GL_INTERFACE data rows.

        Args:
            output_path: Path for the clean template (auto-generated if not provided)

        Returns:
            Path to the clean template file
        """
        if output_path is None:
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            output_path = self.template_path.parent / f"JournalImportTemplate_Clean_{timestamp}.xlsm"
        else:
            output_path = Path(output_path)

        # Create copy and clear data
        shutil.copy2(self.template_path, output_path)
        print(f"Created clean template copy: {output_path}")

        wb = load_workbook(output_path)
        if "GL_INTERFACE" in wb.sheetnames:
            gl_sheet = wb["GL_INTERFACE"]
            # Keep rows 1-4 (instructions and headers), delete data rows from row 5 onwards
            if gl_sheet.max_row > 4:
                gl_sheet.delete_rows(5, gl_sheet.max_row - 4)
                print("Cleared data rows from GL_INTERFACE sheet")
            wb.save(output_path)
            wb.close()

        return str(output_path)

    def fill_template(
        self,
        template_path: str,
        journal_data: List[Dict[str, Any]],
        output_path: Optional[str] = None,
        auto_zip: bool = False,
    ) -> str:
        """
        Fill a template with journal entry data.

        Args:
            template_path: Path to the clean template
            journal_data: List of dictionaries containing journal entry data
            output_path: Path for the filled template (optional)
            auto_zip: Whether to automatically create ZIP file

        Returns:
            Path to the filled template file (or ZIP file if auto_zip=True)
        """
        template_file = Path(template_path)
        if not template_file.exists():
            raise FileNotFoundError(f"Template file not found: {template_path}")
        if not journal_data:
            raise ValueError("No journal data provided")

        if output_path is None:
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            output_path = template_file.parent / f"JournalImport_Filled_{timestamp}.xlsm"
        else:
            output_path = Path(output_path)

        # Create copy and fill with data
        shutil.copy2(template_file, output_path)
        wb = load_workbook(output_path)
        gl_sheet = wb["GL_INTERFACE"]

        # Extract headers from row 4
        headers = []
        for col_num in range(1, gl_sheet.max_column + 1):
            header_cell = gl_sheet.cell(row=4, column=col_num)
            if header_cell.value is not None and str(header_cell.value).strip():
                clean_header = str(header_cell.value).lstrip("*").strip()
                headers.append(clean_header)
            else:
                break

        print(f"Found {len(headers)} headers in template")
        print(f"Filling template with {len(journal_data)} journal entries")

        # Fill data starting from row 5
        for row_idx, entry in enumerate(journal_data, start=5):
            for col_idx, header in enumerate(headers, start=1):
                value = entry.get(header, "")
                cell = gl_sheet.cell(row=row_idx, column=col_idx)
                cell.value = value if value is not None else ""

        wb.save(output_path)
        wb.close()
        print(f"Template filled with data: {output_path}")

        # Optionally create ZIP file
        if auto_zip:
            zip_path = str(output_path).replace(".xlsm", ".zip")
            try:
                zip_result = excel_to_csv_and_zip(str(output_path), zip_path)
                print(f"ZIP file created: {zip_result}")
                return zip_result
            except Exception as e:
                print(f"Warning: Failed to create ZIP file: {e}")
                return str(output_path)

        return str(output_path)

    def create_from_scratch(
        self,
        journal_data: List[Dict[str, Any]],
        output_name: Optional[str] = None,
        auto_zip: bool = True,
    ) -> str:
        """
        Complete workflow: Create clean template, fill with data, and optionally ZIP.

        Args:
            journal_data: List of dictionaries containing journal entry data
            output_name: Base name for output files (optional)
            auto_zip: Whether to create ZIP file automatically

        Returns:
            Path to the final file (Excel or ZIP)
        """
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        if output_name is None:
            output_name = f"JournalImport_{timestamp}"

        try:
            # Create clean template
            clean_template = self.create_clean_template()

            # Fill with data
            filled_template_path = self.template_path.parent / f"{output_name}.xlsm"
            result_path = self.fill_template(
                clean_template, journal_data, str(filled_template_path), auto_zip=auto_zip
            )

            # Clean up temporary clean template
            Path(clean_template).unlink(missing_ok=True)

            return result_path

        except Exception as e:
            print(f"Error in journal creation workflow: {e}")
            raise


def create_journal_entry_data(
    transfers,
    transaction_id: int = 0,
    entry_type: str = "submit",
    group_id: int = 0
) -> List[Dict[str, Any]]:
    """
    Create journal entry data with dynamic segment support from transfer objects.

    Args:
        transfers: List of TransactionTransfer objects with segment relationships
        transaction_id: Transaction identifier
        entry_type: "submit" or "reject" - determines debit/credit direction
        group_id: Interface group identifier for Oracle

    Returns:
        List of journal entries with dynamic SEGMENT1-SEGMENT30 columns
    """
    mapper = OracleSegmentMapper()
    
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    # batch_name = f"BATCH_TRANSFER_{timestamp}_TXN_{transaction_id}"
    batch_description = f"Balance Transfer Batch - Transaction {transaction_id}"
    journal_name = f"JOURNAL_TRANSFER_{timestamp}_TXN_{transaction_id}"
    journal_description = f"Journal Entry for Balance Transfer - Transaction {transaction_id}"

    # Load configuration from environment
    CURRENCY_CODE = os.getenv("ORACLE_CURRENCY_CODE", "AED")
    LEDGER_ID = os.getenv("ORACLE_LEDGER_ID", "300000205309206")
    ORACLE_EFFECTIVE_DATE = os.getenv("ORACLE_EFFECTIVE_DATE", "2025-09-26")
    ORACLE_JOURNAL_SOURCE = os.getenv("ORACLE_JOURNAL_SOURCE", "Allocations")
    ENCUMBRANCE_TYPE_ID = os.getenv("ENCUMBRANCE_TYPE_ID", "300000035858125")
    
    sample_data = []
    total_debit = 0
    
    budget_trasnfer=xx_BudgetTransfer.objects.get(transaction_id=transaction_id)
    linked_transfer=budget_trasnfer.linked_transfer_id
    
    # ========== HFR REJECTION SPECIAL LOGIC ==========
    # For HFR rejection, only return the REMAINING unused amount
    hfr_remaining_amount = None
    if entry_type == "reject" and budget_trasnfer.code and budget_trasnfer.code[0:3] == "HFR":
        from decimal import Decimal
        
        # Calculate original hold amount
        hfr_transfers = list(transfers)
        original_hold = sum(
            Decimal(str(t.from_center)) if t.from_center else Decimal('0.00')
            for t in hfr_transfers
        )
        
        # Find all approved/in-progress FAR transfers linked to this HFR
        from django.db.models import Q
        linked_fars = xx_BudgetTransfer.objects.filter(
            linked_transfer_id=transaction_id,
            code__startswith="FAR"
        ).filter(
            Q(status="approved") | Q(status_level__gte=2)
        ).exclude(
            status_level__lt=1
        )
        
        # Calculate total used
        total_used = Decimal('0.00')
        for far in linked_fars:
            from transaction.models import xx_TransactionTransfer as TT
            far_transfers = TT.objects.filter(transaction_id=far.transaction_id)
            far_total = sum(
                Decimal(str(t.from_center)) if t.from_center else Decimal('0.00')
                for t in far_transfers
            )
            total_used += far_total
        
        # Calculate remaining
        hfr_remaining_amount = original_hold - total_used
        
        print(f"🔴 HFR REJECTION Journal Entry:")
        print(f"   Original Hold: {original_hold}")
        print(f"   Already Used: {total_used}")
        print(f"   Remaining to Return: {hfr_remaining_amount}")
        
        # If nothing remaining, return empty journal (no entry needed)
        if hfr_remaining_amount <= 0:
            print(f"   ⚠️  No remaining amount - skipping journal entry")
            return sample_data, False, False, True  # Return empty with status=False
    
    # Calculate total for offsetting entry
    for transfer in transfers:
        total_debit += getattr(transfer, "from_center", 0) or 0


    # Create journal lines for each transfer (FROM side - debit entries)
    journal_catgorays=os.getenv("ORACLE_journal_catgorays")
    journal_catgorays=journal_catgorays.split(",")
    print(f"journal_catgorays: {journal_catgorays}")
    journal_catgoray=""
    if budget_trasnfer.control_budget=="سيولة":
        journal_catgoray=journal_catgorays[0]
    elif budget_trasnfer.control_budget=="تكاليف":
        journal_catgoray=journal_catgorays[1]

    print(f"journal_catgoray: {journal_catgoray}")

    batch_name = f"BATCH_{journal_catgoray}_{timestamp}_TXN_{transaction_id}"
    
    # Pre-fetch linked transfer segments and amounts if linked_transfer exists
    linked_transfer_data = {}
    if linked_transfer is not None and budget_trasnfer.code[0:3] == "FAR":
        from account_and_entitys.models import XX_TransactionSegment
        linked_transfers_list = xx_TransactionTransfer.objects.filter(transaction=linked_transfer)
        for linked_trans in linked_transfers_list:
            # Get all segments for this linked transfer
            segments = XX_TransactionSegment.objects.filter(
                transaction_transfer=linked_trans
            ).select_related('segment_type', 'from_segment_value', 'to_segment_value')
            
            # Build segment key (tuple of all segment codes) to match transfers
            segment_key = tuple(
                (seg.segment_type_id, seg.from_segment_value.code if seg.from_segment_value else None)
                for seg in segments.order_by('segment_type_id')
            )
            
            linked_transfer_data[segment_key] = {
                'from_center': linked_trans.from_center,
                'to_center': linked_trans.to_center,
                'transfer': linked_trans
            }
    
    for transfer in transfers:
        # Check if this transfer has a matching linked transfer
        linked_from_amount = 0
        current_from_amount = getattr(transfer, "from_center", 0) or 0
        has_linked_match = False
        
        if linked_transfer is not None and linked_transfer_data:
            from account_and_entitys.models import XX_TransactionSegment
            current_segments = XX_TransactionSegment.objects.filter(
                transaction_transfer=transfer
            ).select_related('segment_type', 'from_segment_value', 'to_segment_value')
            
            # Build segment key for current transfer
            current_segment_key = tuple(
                (seg.segment_type_id, seg.from_segment_value.code if seg.from_segment_value else None)
                for seg in current_segments.order_by('segment_type_id')
            )
            
            # Find matching linked transfer by segment combination
            if current_segment_key in linked_transfer_data:
                linked_data = linked_transfer_data[current_segment_key]
                linked_from_amount = linked_data['from_center'] or 0
                has_linked_match = True
                
                print(f"Transfer {transfer.transfer_id} - Current: {current_from_amount}, Linked: {linked_from_amount}")
        
        # Determine debit and credit amounts based on linked transfer
        if has_linked_match:
            # DEBIT is always the current amount
            debit_amount = current_from_amount
            
            # CREDIT is the minimum of current and linked (the amount that matches/cancels)
            credit_amount = min(current_from_amount, linked_from_amount)
            
            # Adjust total_debit: subtract the credit (which cancels part of debit)
            # Net effect = current - min(current, linked)
            total_debit = total_debit - credit_amount
        else:
            # No linked match - use full amount as debit
            debit_amount = current_from_amount
            credit_amount = 0
        
        # ========== HFR REJECTION: Use proportional remaining amount ==========
        if hfr_remaining_amount is not None and entry_type == "reject":
            # For HFR rejection, proportionally reduce each transfer's amount
            # based on the ratio: remaining / original_hold
            from decimal import Decimal
            
            # Calculate what percentage of original we're returning
            original_transfer_amount = Decimal(str(current_from_amount))
            
            # Get total original hold (sum of all from_center values)
            total_original = sum(
                Decimal(str(t.from_center)) if t.from_center else Decimal('0.00')
                for t in transfers
            )
            
            if total_original > 0:
                # Calculate proportional amount for this transfer
                proportion = original_transfer_amount / total_original
                proportional_amount = float(hfr_remaining_amount * proportion)
                
                debit_amount = proportional_amount
                credit_amount = 0
                
                print(f"   📝 Transfer {transfer.transfer_id}: Original={original_transfer_amount}, "
                      f"Proportion={proportion:.2%}, Remaining Amount={proportional_amount:.2f}")
            else:
                # Edge case: total is 0, skip this transfer
                print(f"   ⏭️  Skipping transfer {transfer.transfer_id} (total original is 0)")
                continue
        
        if debit_amount > 0:
            # Create DEBIT entry
            journal_entry = {
                "Status Code": "NEW",
                "Ledger ID": LEDGER_ID,
                "Effective Date of Transaction": ORACLE_EFFECTIVE_DATE,
                "Journal Source": ORACLE_JOURNAL_SOURCE,
                "Journal Category": journal_catgoray,
                "Currency Code": CURRENCY_CODE,
                "Journal Entry Creation Date": time.strftime("%Y-%m-%d"),
                "Actual Flag": "E",
                "Entered Debit Amount": debit_amount if entry_type == "submit" else "",
                "Entered Credit Amount": debit_amount if entry_type == "reject" else "",
                "REFERENCE1 (Batch Name)": batch_name,
                "REFERENCE2 (Batch Description)": batch_description,
                "REFERENCE4 (Journal Entry Name)": journal_name,
                "REFERENCE5 (Journal Entry Description)": journal_description,
                "REFERENCE10 (Journal Entry Line Description)": f"Debit transaction {transfer.transaction_id} transfer line {transfer.transfer_id}",
                "Encumbrance Type ID": ENCUMBRANCE_TYPE_ID,
                "Interface Group Identifier": group_id,
            }
            segment_data = mapper.build_fbdi_row(
                transaction_transfer=transfer,
                base_row=journal_entry,
                fill_all=True
            )
            sample_data.append(segment_data)
        
        # If there's a linked match, create CREDIT entry for the same segment
        if has_linked_match and credit_amount > 0:
            credit_entry = {
                "Status Code": "NEW",
                "Ledger ID": LEDGER_ID,
                "Effective Date of Transaction": ORACLE_EFFECTIVE_DATE,
                "Journal Source": ORACLE_JOURNAL_SOURCE,
                "Journal Category": journal_catgoray,
                "Currency Code": CURRENCY_CODE,
                "Journal Entry Creation Date": time.strftime("%Y-%m-%d"),
                "Actual Flag": "E",
                "Entered Debit Amount": credit_amount if entry_type == "reject" else "",
                "Entered Credit Amount": credit_amount if entry_type == "submit" else "",
                "REFERENCE1 (Batch Name)": batch_name,
                "REFERENCE2 (Batch Description)": batch_description,
                "REFERENCE4 (Journal Entry Name)": journal_name,
                "REFERENCE5 (Journal Entry Description)": journal_description,
                "REFERENCE10 (Journal Entry Line Description)": f"Credit for linked transfer - transaction {transfer.transaction_id} transfer line {transfer.transfer_id}",
                "Encumbrance Type ID": ENCUMBRANCE_TYPE_ID,
                "Interface Group Identifier": group_id,
            }
            # Use same segments as the debit entry
            credit_segment_data = mapper.build_fbdi_row(
                transaction_transfer=transfer,
                base_row=credit_entry,
                fill_all=True
            )
            sample_data.append(credit_segment_data)

    # ========== HFR REJECTION: Use remaining amount for offsetting entry ==========
    if hfr_remaining_amount is not None and entry_type == "reject":
        total_debit = float(hfr_remaining_amount)
        print(f"   📊 Setting offsetting entry to remaining amount: {total_debit}")

    if total_debit > 0 and transfers:
        
        # Static segment values for offsetting entry
        offsetting_entry = {
            "Status Code": "NEW",
            "Ledger ID": LEDGER_ID,
            "Effective Date of Transaction": ORACLE_EFFECTIVE_DATE,
            "Journal Source": ORACLE_JOURNAL_SOURCE,
            "Journal Category": journal_catgoray,
            "Currency Code": CURRENCY_CODE,
            "Journal Entry Creation Date": time.strftime("%Y-%m-%d"),
            "Actual Flag": "E",
            "Entered Debit Amount": total_debit if entry_type == "reject" else "",
            "Entered Credit Amount": total_debit if entry_type == "submit" else "",
            "REFERENCE1 (Batch Name)": batch_name,
            "REFERENCE2 (Batch Description)": batch_description,
            "REFERENCE4 (Journal Entry Name)": journal_name,
            "REFERENCE5 (Journal Entry Description)": journal_description,
            "REFERENCE10 (Journal Entry Line Description)": "Offsetting credit line for balance transfer",
            "Encumbrance Type ID": ENCUMBRANCE_TYPE_ID,
            "Interface Group Identifier": group_id,
            # Static segment values
            "Segment1": "1",
            "Segment2": "0138",
            "Segment3": "1",
            "Segment4": "00000000",
            "Segment5": "6829999",
            "Segment6": "412119997",
            "Segment7": "00000",
            "Segment8": "70113",
            "Segment9": "013810010001",
            "Segment10": "0000",
            "Segment11": "000000000000",
            "Segment12": "00000",
            "Segment13": "00000",
            "Segment14": "00000",
            "Segment15": "00000",
            "Segment16": "00000",
        }
        
        # Fill remaining segments (17-30) with empty values
        for i in range(17, 31):
            offsetting_entry[f"Segment{i}"] = ""
        
        offsetting_with_segments = offsetting_entry
        sample_data.append(offsetting_with_segments)

    return sample_data, True 


def create_custom_journal_entry(
    segment_values: Dict[int, str],
    ledger_id: str = "300000205309206",
    effective_date: Optional[str] = None,
    debit_amount: float = 0,
    credit_amount: float = 0,
    batch_name: Optional[str] = None,
    journal_name: Optional[str] = None,
    line_description: str = "",
    group_id: int = 0,
    **additional_fields
) -> Dict[str, Any]:
    """
    Create a single journal entry with custom segment values.

    Args:
        segment_values: Dict mapping segment_type_id to segment code (e.g., {1: 'E001', 2: 'A100'})
        ledger_id: Oracle ledger ID
        effective_date: Transaction date (YYYY-MM-DD format)
        debit_amount: Debit amount (0 if credit entry)
        credit_amount: Credit amount (0 if debit entry)
        batch_name: Batch identifier
        journal_name: Journal entry identifier
        line_description: Description for this journal line
        group_id: Interface group identifier
        **additional_fields: Any other FBDI columns

    Returns:
        Dictionary with all FBDI columns including dynamic SEGMENT1-SEGMENT30
    """
    mapper = OracleSegmentMapper()
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    
    if effective_date is None:
        effective_date = time.strftime("%Y-%m-%d")
    if batch_name is None:
        batch_name = f"BATCH_{timestamp}"
    if journal_name is None:
        journal_name = f"JOURNAL_{timestamp}"
    
    base_entry = {
        "Status Code": "NEW",
        "Ledger ID": ledger_id,
        "Effective Date of Transaction": effective_date,
        "Journal Source": "Allocations",
        "Journal Category": "Adjustment",
        "Currency Code": "AED",
        "Journal Entry Creation Date": time.strftime("%Y-%m-%d"),
        "Actual Flag": "E",
        "Entered Debit Amount": debit_amount if debit_amount > 0 else "",
        "Entered Credit Amount": credit_amount if credit_amount > 0 else "",
        "REFERENCE1 (Batch Name)": batch_name,
        "REFERENCE2 (Batch Description)": f"Batch created {time.strftime('%Y-%m-%d %H:%M:%S')}",
        "REFERENCE4 (Journal Entry Name)": journal_name,
        "REFERENCE5 (Journal Entry Description)": f"Journal entry {time.strftime('%Y-%m-%d %H:%M:%S')}",
        "REFERENCE10 (Journal Entry Line Description)": line_description,
        "Encumbrance Type ID": "100000243328511",
        "Interface Group Identifier": group_id,
    }
    
    base_entry.update(additional_fields)
    
    # Build segment columns dynamically
    for segment_type_id, segment_code in segment_values.items():
        oracle_field_name = mapper.get_oracle_field_name(segment_type_id)
        if oracle_field_name:
            base_entry[oracle_field_name] = segment_code
    
    # Fill remaining segments
    configured_segments = mapper.get_active_oracle_fields()
    all_segments = mapper.build_all_oracle_fields(num_segments=30)
    
    for segment_field in all_segments:
        if segment_field not in base_entry:
            base_entry[segment_field] = "00000" if segment_field not in configured_segments else ""
    
    return base_entry


def create_balanced_journal_pair(
    debit_segments: Dict[int, str],
    credit_segments: Dict[int, str],
    amount: float,
    ledger_id: str = "300000205309206",
    effective_date: Optional[str] = None,
    batch_name: Optional[str] = None,
    journal_name: Optional[str] = None,
    debit_description: str = "Debit entry",
    credit_description: str = "Credit entry",
    group_id: int = 0,
) -> List[Dict[str, Any]]:
    """
    Create a balanced pair of journal entries (debit + credit).

    Args:
        debit_segments: Segment values for debit entry
        credit_segments: Segment values for credit entry
        amount: Transaction amount (positive value)
        ledger_id: Oracle ledger ID
        effective_date: Transaction date
        batch_name: Batch identifier (auto-generated if None)
        journal_name: Journal identifier (auto-generated if None)
        debit_description: Description for debit line
        credit_description: Description for credit line
        group_id: Interface group identifier

    Returns:
        List of 2 journal entries (debit and credit)
    """
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    
    if batch_name is None:
        batch_name = f"BATCH_{timestamp}"
    if journal_name is None:
        journal_name = f"JOURNAL_{timestamp}"
    
    debit_entry = create_custom_journal_entry(
        segment_values=debit_segments,
        ledger_id=ledger_id,
        effective_date=effective_date,
        debit_amount=amount,
        credit_amount=0,
        batch_name=batch_name,
        journal_name=journal_name,
        line_description=debit_description,
        group_id=group_id,
    )
    
    credit_entry = create_custom_journal_entry(
        segment_values=credit_segments,
        ledger_id=ledger_id,
        effective_date=effective_date,
        debit_amount=0,
        credit_amount=amount,
        batch_name=batch_name,
        journal_name=journal_name,
        line_description=credit_description,
        group_id=group_id,
    )
    
    return [debit_entry, credit_entry]
