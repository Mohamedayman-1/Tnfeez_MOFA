from __future__ import annotations

import html as _html
import re
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple, Union


def _peek_head_bytes(path: Path, size: int = 512) -> bytes:
    try:
        return path.read_bytes()[:size]
    except OSError:
        return b""


def _looks_like_html(head: bytes) -> bool:
    if not head:
        return False
    sample = head.lstrip().lower()
    return sample.startswith((b"<html", b"<!doctype html")) or b"<table" in sample[:128]


def _normalize_headers(raw_headers: Iterable[Any]) -> List[str]:
    normalized: List[str] = []
    seen = set()
    for idx, value in enumerate(raw_headers):
        name = ""
        if value is not None:
            name = str(value).strip()
        if not name:
            name = f"Column{idx+1}"
        base = name
        suffix = 1
        while name in seen:
            suffix += 1
            name = f"{base}_{suffix}"
        seen.add(name)
        normalized.append(name)
    return normalized


def _rows_to_records(rows_iter: Iterable[Iterable[Any]], headers: List[str], fillna: Any) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    header_len = len(headers)
    for row in rows_iter:
        values = list(row)
        if len(values) < header_len:
            values += [None] * (header_len - len(values))
        elif len(values) > header_len:
            values = values[:header_len]
        if fillna is not None:
            values = [fillna if v is None else v for v in values]
        records.append(dict(zip(headers, values)))
    return records


def _prepend_first(first_row: Iterable[Any], rows_iter: Iterable[Iterable[Any]]) -> Iterable[Iterable[Any]]:
    yield first_row
    for row in rows_iter:
        yield row


CLAUSE_PATTERN = re.compile(
    r"(?P<field>\w+)\s+(?P<op>(?i:equals to|not equals to|between|not between))\s+(?P<value>.*?)(?=\s+(?:OR|AND)\s+|$)",
)

AND_SPLIT_RE = re.compile(r"\s+and\s+", re.IGNORECASE)

class _HTMLTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.tables: List[List[List[str]]] = []
        self._table_depth = 0
        self._collecting = False
        self._current_table: List[List[str]] = []
        self._current_row: Optional[List[str]] = None
        self._current_cell_chunks: List[str] = []
        self._capturing_cell = False

    def handle_starttag(self, tag: str, attrs: List[tuple]):
        tag = tag.lower()
        if tag == "table":
            self._table_depth += 1
            if self._table_depth == 1 and not self._collecting:
                self._collecting = True
                self._current_table = []
        if self._collecting and self._table_depth >= 1:
            if tag == "tr":
                self._current_row = []
            elif tag in ("td", "th"):
                self._capturing_cell = True
                self._current_cell_chunks = []
            elif tag == "br" and self._capturing_cell:
                self._current_cell_chunks.append("\n")

    def handle_endtag(self, tag: str):
        tag = tag.lower()
        if self._collecting:
            if tag in ("td", "th") and self._capturing_cell:
                text = "".join(self._current_cell_chunks)
                text = _html.unescape(text).replace("\xa0", " ").strip()
                if self._current_row is None:
                    self._current_row = []
                self._current_row.append(text)
                self._capturing_cell = False
                self._current_cell_chunks = []
            elif tag == "tr" and self._current_row is not None:
                self._current_table.append(self._current_row)
                self._current_row = None
        if tag == "table":
            if self._collecting and self._table_depth == 1:
                if self._current_row is not None:
                    self._current_table.append(self._current_row)
                    self._current_row = None
                if self._current_table:
                    self.tables.append(self._current_table)
                self._collecting = False
                self._current_table = []
            if self._table_depth > 0:
                self._table_depth -= 1

    def handle_data(self, data: str):
        if self._capturing_cell:
            self._current_cell_chunks.append(data)

    def handle_startendtag(self, tag: str, attrs: List[tuple]):
        if tag.lower() == "br" and self._capturing_cell:
            self._current_cell_chunks.append("\n")

    def first_table(self) -> List[List[str]]:
        return self.tables[0] if self.tables else []


def _parse_html_table(path: Path, header: bool, fillna: Any) -> List[Dict[str, Any]]:
    text = path.read_text(encoding="utf-8", errors="ignore")
    parser = _HTMLTableParser()
    parser.feed(text)
    parser.close()
    table = parser.first_table()
    if not table:
        return []

    rows = [list(row) for row in table]
    if not rows:
        return []

    if header:
        headers = _normalize_headers(rows[0])
        data_rows = rows[1:]
    else:
        width = max((len(row) for row in rows), default=0)
        headers = [f"Column{i+1}" for i in range(width)]
        data_rows = rows

    records: List[Dict[str, Any]] = []
    for row in data_rows:
        values: List[Any] = []
        for idx in range(len(headers)):
            value: Any = row[idx] if idx < len(row) else None
            if isinstance(value, str):
                cleaned = value.strip()
                value = cleaned if cleaned else None
            if fillna is not None and value is None:
                value = fillna
            values.append(value)
        records.append(dict(zip(headers, values)))

    return records


def _parse_rule_clauses(text: Optional[str]) -> List[Dict[str, Union[str, Tuple[str, str]]]]:
    """Split a rule cell into structured clauses."""

    if not text:
        return []

    clauses: List[Dict[str, Union[str, Tuple[str, str]]]] = []
    for match in CLAUSE_PATTERN.finditer(text):
        field = match.group("field")
        op_token = match.group("op").lower().replace(" ", "_")
        value_raw = match.group("value").strip()

        if op_token in {"between", "not_between"}:
            parts = [part.strip() for part in AND_SPLIT_RE.split(value_raw) if part.strip()]
            if len(parts) >= 2:
                value = (parts[0], parts[1])
            else:
                value = value_raw
        else:
            value = value_raw

        clauses.append({"field": field, "op": op_token, "value": value})

    return clauses


def read_excel_to_dicts(
    file_path: Union[str, Path],
    sheet: Optional[Union[int, str]] = 0,
    header: bool = True,
    fillna: Any = None,
) -> List[Dict[str, Any]]:
    """
    Read data from an Excel file and return it as a list of row dictionaries.

    - Uses pandas if available for robustness and speed.
    - Falls back to openpyxl without pandas.

    Args:
        file_path: Path to the Excel file (e.g., .xlsx, .xls).
        sheet: Sheet index (int) or name (str). Defaults to first sheet.
        header: Whether the first row is a header row. If False, columns are
                named as Column1, Column2, ...
        fillna: Value to replace empty cells (None). If left as None, values
                are returned as-is (including None for empty cells).

    Returns:
        List of dictionaries, each dict representing a row (keyed by column name).

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the specified sheet cannot be found.
    """

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    ext = path.suffix.lower()
    head_bytes = _peek_head_bytes(path)

    if ext == ".xls" and _looks_like_html(head_bytes):
        return _parse_html_table(path, header, fillna)

    # Try pandas first if available
    try:
        import pandas as pd  # type: ignore

        sheet_name = sheet if sheet is not None else 0
        engine_map = {
            ".xlsx": "openpyxl",
            ".xlsm": "openpyxl",
            ".xltx": "openpyxl",
            ".xltm": "openpyxl",
            ".xls": "xlrd",
            ".xlsb": "pyxlsb",
        }
        engine = engine_map.get(ext)

        df = pd.read_excel(
            path,
            sheet_name=sheet_name,
            dtype=object,
            header=0 if header else None,
            engine=engine,
        )

        if not header:
            df.columns = [f"Column{i+1}" for i in range(len(df.columns))]

        if fillna is not None:
            df = df.fillna(fillna)

        return df.to_dict(orient="records")

    except (ModuleNotFoundError, ImportError, ValueError):
        pass

    if ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ModuleNotFoundError as e:
            raise ModuleNotFoundError(
                "Reading Excel files requires 'openpyxl' when pandas is unavailable."
            ) from e

        wb = load_workbook(filename=str(path), data_only=True, read_only=True)

        if not wb.worksheets:
            return []

        ws = None
        if sheet is None or sheet == 0:
            ws = wb.worksheets[0]
        elif isinstance(sheet, int):
            try:
                ws = wb.worksheets[sheet]
            except IndexError as e:
                raise ValueError(f"Sheet index out of range: {sheet}") from e
        elif isinstance(sheet, str):
            if sheet not in wb.sheetnames:
                raise ValueError(f"Sheet not found: {sheet}")
            ws = wb[sheet]
        else:
            raise TypeError("Sheet must be specified by index or name.")

        if ws is None:
            raise ValueError("No sheets found in the workbook.")

        rows_iter = ws.iter_rows(values_only=True)
        try:
            first_row = tuple(next(rows_iter))
        except StopIteration:
            return []

        if header:
            headers = _normalize_headers(first_row)
        else:
            headers = [f"Column{i+1}" for i in range(len(first_row))]
            rows_iter = _prepend_first(first_row, rows_iter)

        return _rows_to_records(rows_iter, headers, fillna)

    if ext == ".xls":
        try:
            import xlrd  # type: ignore
        except ModuleNotFoundError as e:
            raise ModuleNotFoundError(
                "Reading .xls requires 'xlrd'. Install it via 'pip install xlrd'."
            ) from e

        book = xlrd.open_workbook(str(path), on_demand=True)

        if sheet is None or sheet == 0:
            sh = book.sheet_by_index(0)
        elif isinstance(sheet, int):
            try:
                sh = book.sheet_by_index(sheet)
            except IndexError as e:
                raise ValueError(f"Sheet index out of range: {sheet}") from e
        elif isinstance(sheet, str):
            try:
                sh = book.sheet_by_name(sheet)
            except xlrd.biffh.XLRDError as e:  # type: ignore[attr-defined]
                raise ValueError(f"Sheet not found: {sheet}") from e
        else:
            raise TypeError("Sheet must be specified by index or name.")

        if sh is None:
            raise ValueError("No sheets found in the workbook.")

        nrows = sh.nrows
        if nrows == 0:
            return []

        first_row = sh.row_values(0)
        if header:
            headers = _normalize_headers(first_row)
            start_idx = 1
        else:
            headers = [f"Column{i+1}" for i in range(len(first_row))]
            start_idx = 0

        records: List[Dict[str, Any]] = []
        for r in range(start_idx, nrows):
            values = sh.row_values(r)
            if len(values) < len(headers):
                values += [None] * (len(headers) - len(values))
            elif len(values) > len(headers):
                values = values[: len(headers)]

            cleaned: List[Any] = []
            for value in values:
                if isinstance(value, str):
                    stripped = value.strip()
                    value = stripped if stripped else None
                elif value == "":
                    value = None
                if fillna is not None and value is None:
                    value = fillna
                cleaned.append(value)
            records.append(dict(zip(headers, cleaned)))

        return records

    if ext == ".xlsb":
        raise ModuleNotFoundError(
            "Reading .xlsb files requires pandas with the 'pyxlsb' engine installed."
        )

    raise ValueError(
        f"Unsupported Excel extension '{ext}'. Supported: .xlsx, .xlsm, .xltx, .xltm, .xls, .xlsb"
    )


def export_rules_to_excel(
    source_path: Union[str, Path],
    target_path: Union[str, Path],
) -> None:
    """Convert cross-validation rules into a normalized Excel sheet."""

    try:
        from openpyxl import Workbook  # type: ignore
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "export_rules_to_excel requires 'openpyxl'. Install it via 'pip install openpyxl'."
        ) from exc

    data = read_excel_to_dicts(source_path)

    op_display = {
        "equals_to": "=",
        "not_equals_to": "!=",
        "between": "Between",
        "not_between": "not Between",
    }

    header = [
        "Condition Segment",
        "Condition Operation",
        "Condition Value",
        "Validation Segment",
        "Validation Operation",
        "Validation Value",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Normalized Rules"
    ws.append(header)

    for row in data:
        conditions = _parse_rule_clauses(row.get("Condition Details"))
        validations = _parse_rule_clauses(row.get("Validation Details"))

        if not conditions:
            conditions = [None]
        if not validations:
            validations = [None]

        for cond in conditions:
            cond_segment = cond["field"] if isinstance(cond, dict) else ""
            cond_op_key = cond["op"] if isinstance(cond, dict) else ""
            cond_value = cond["value"] if isinstance(cond, dict) else ""

            if isinstance(cond_value, tuple):
                cond_value = ",".join(cond_value)
            elif cond_value is None:
                cond_value = ""
            else:
                cond_value = str(cond_value)

            cond_op = op_display.get(cond_op_key, cond_op_key) if cond_op_key else ""

            for val in validations:
                val_segment = val["field"] if isinstance(val, dict) else ""
                val_op_key = val["op"] if isinstance(val, dict) else ""
                val_value = val["value"] if isinstance(val, dict) else ""

                if isinstance(val_value, tuple):
                    val_value = ",".join(val_value)
                elif val_value is None:
                    val_value = ""
                else:
                    val_value = str(val_value)

                val_op = op_display.get(val_op_key, val_op_key) if val_op_key else ""

                ws.append([
                    cond_segment,
                    cond_op,
                    cond_value,
                    val_segment,
                    val_op,
                    val_value,
                ])

    wb.save(str(target_path))


__all__ = ["read_excel_to_dicts", "export_rules_to_excel"]

export_rules_to_excel("Cross Validation.xls", "Normalized_Rules.xlsx")



# def Check_Validation_Rules(Entity, Account, Project):


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Read cross-validation rules or export them to a normalized Excel file.",
    )
    parser.add_argument(
        "source",
        nargs="?",
        default="Cross Validation.xls",
        help="Path to the source Excel file (default: %(default)s)",
    )
    parser.add_argument(
        "target",
        nargs="?",
        help="Path to write the normalized Excel file. If omitted, just prints the row count.",
    )
    args = parser.parse_args()

    if args.target:
        export_rules_to_excel(args.source, args.target)
        print(f"Exported normalized rules to {args.target}")
    else:
        rows = read_excel_to_dicts(args.source)
        print(f"Loaded {len(rows)} rows from {args.source}")
