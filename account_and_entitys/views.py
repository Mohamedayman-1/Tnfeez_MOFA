import numpy as np
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination

from budget_management.models import (
    get_entities_with_children,
    get_level_zero_children,
    get_zero_level_accounts,
    get_zero_level_projects,
)
from .models import (
    XX_Account,
    XX_Entity,
    XX_Project,
    XX_PivotFund,
    XX_TransactionAudit,
    XX_ACCOUNT_ENTITY_LIMIT,
    XX_BalanceReport,
    Account_Mapping,
    Budget_data,
    XX_ACCOUNT_mapping,
    XX_Entity_mapping,
    EnvelopeManager,
    XX_Segment,
    XX_SegmentType,
)
from .serializers import (
    AccountSerializer,
    EntitySerializer,
    PivotFundSerializer,
    ProjectSerializer,
    TransactionAuditSerializer,
    AccountEntityLimitSerializer,
    BalanceReportSerializer,
    SegmentValueListSerializer,
)
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework import status
import pandas as pd
from django.db import transaction
from .models import XX_ACCOUNT_ENTITY_LIMIT
from .serializers import AccountEntityLimitSerializer
from django.db.models import CharField
from django.db.models.functions import Cast
from django.db.models import Q
from .utils import get_oracle_report_data, get_mapping_for_fusion_data
from .oracle.oracle_balance_report_manager import OracleBalanceReportManager


class EntityPagination(PageNumberPagination):
    """Pagination class for entities and accounts"""

    page_size = 10
    page_size_query_param = "page_size"
    max_page_size = 100


# Unified Dynamic Segment View

################################################################################# Segment types are now stored in XX_SegmentType table##############################################################
class SegmentTypesListView(APIView):
    """List all available segment types in the system
    
    Use this endpoint to discover what segment types are configured and available.
    Returns all segment types from XX_SegmentType table with their counts.
    """
    
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        from django.db.models import Count
        
        # Get all segment types with counts
        segment_types = XX_SegmentType.objects.all().order_by('segment_id')
        
        types_data = []
        for seg_type in segment_types:
            segment_count = XX_Segment.objects.filter(
                segment_type=seg_type,
                is_active=True
            ).count()
            
            types_data.append({
                "segment_id": seg_type.segment_id,
                "segment_name": seg_type.segment_name,
                "description": seg_type.description if hasattr(seg_type, 'description') else None,
                "total_segments": segment_count,
                "endpoint_example": f"/api/accounts-entities/segments/?segment_type={seg_type.segment_id}",
                "endpoint_by_name": f"/api/accounts-entities/segments/?segment_type={seg_type.segment_name}"
            })
        
        return Response(
            {
                "message": "Available segment types retrieved successfully.",
                "total_types": len(types_data),
                "data": types_data
            }
        )


class SegmentTypeCreateView(APIView):
    """Create a new segment type
    
    Request Body:
    - segment_name: REQUIRED - Name of the segment type (e.g., "Department", "Location")
    - description: OPTIONAL - Description of this segment type
    
    Note: segment_id is auto-generated
    """
    
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        segment_name = request.data.get("segment_name")
        oracle_segment_number=request.data.get("oracle_segment_number")
        is_required=request.data.get("is_required", True)
        has_hierarchy=request.data.get("has_hierarchy", False)
        display_order=request.data.get("display_order", 0)
        description=request.data.get("description")
        if not segment_name:
            return Response(
                {
                    "message": "segment_name is required in request body.",
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if segment type with this name already exists
        existing = XX_SegmentType.objects.filter(
            segment_name__iexact=segment_name
        ).first()
        
        if existing:
            return Response(
                {
                    "message": f"Segment type with name '{segment_name}' already exists.",
                    "existing_segment_type": {
                        "segment_id": existing.segment_id,
                        "segment_name": existing.segment_name,
                        "description": existing.description if hasattr(existing, 'description') else None
                    }
                },
                status=status.HTTP_409_CONFLICT
            )
        
        
        try:
            segment_type = XX_SegmentType.objects.create(
                segment_name=segment_name,
                description=description,
                oracle_segment_number=oracle_segment_number,
                is_required=is_required,
                has_hierarchy=has_hierarchy,
                display_order=display_order
            )
            
            return Response(
                {
                    "message": f"Segment type '{segment_name}' created successfully.",
                    "data": {
                        "segment_id": segment_type.segment_id,
                        "segment_name": segment_type.segment_name,
                        "description": segment_type.description if hasattr(segment_type, 'description') else None
                    }
                },
                status=status.HTTP_201_CREATED
            )
        
        except Exception as e:
            return Response(
                {
                    "message": "Failed to create segment type.",
                    "error": str(e)
                },
                status=status.HTTP_400_BAD_REQUEST
            )


class SegmentTypeDetailView(APIView):
    """Retrieve a specific segment type by ID"""
    
    permission_classes = [IsAuthenticated]
    
    def get(self, request, pk):
        try:
            segment_type = XX_SegmentType.objects.get(segment_id=pk)
        except XX_SegmentType.DoesNotExist:
            return Response(
                {"message": "Segment type not found."},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Get count of segments for this type
        segment_count = XX_Segment.objects.filter(
            segment_type=segment_type,
            is_active=True
        ).count()
        
        return Response(
            {
                "message": "Segment type details retrieved successfully.",
                "data": {
                    "segment_id": segment_type.segment_id,
                    "segment_name": segment_type.segment_name,
                    "description": segment_type.description if hasattr(segment_type, 'description') else None,
                    "total_segments": segment_count,
                    "endpoint_example": f"/api/accounts-entities/segments/?segment_type={segment_type.segment_id}",
                    "endpoint_by_name": f"/api/accounts-entities/segments/?segment_type={segment_type.segment_name}"
                }
            }
        )


class SegmentTypeUpdateView(APIView):
    """Update a specific segment type
    
    Supports updating:
    - segment_name (must remain unique)
    - description
    
    Note: segment_id cannot be changed as it's the primary key
    """
    
    permission_classes = [IsAuthenticated]
    
    def put(self, request, pk):
        try:
            segment_type = XX_SegmentType.objects.get(segment_id=pk)
        except XX_SegmentType.DoesNotExist:
            return Response(
                {"message": "Segment type not found."},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Check if segment_name is being changed and if it conflicts
        new_name = request.data.get("segment_name")
        if new_name and new_name != segment_type.segment_name:
            # Check for duplicate name
            existing = XX_SegmentType.objects.filter(
                segment_name__iexact=new_name
            ).exclude(segment_id=pk).first()
            
            if existing:
                return Response(
                    {
                        "message": f"Segment type with name '{new_name}' already exists.",
                    },
                    status=status.HTTP_409_CONFLICT
                )
            
            segment_type.segment_name = new_name
        
        # Update description if provided
        if "description" in request.data:
            segment_type.description = request.data.get("description")
        
        try:
            segment_type.save()
            
            return Response(
                {
                    "message": "Segment type updated successfully.",
                    "data": {
                        "segment_id": segment_type.segment_id,
                        "segment_name": segment_type.segment_name,
                        "description": segment_type.description if hasattr(segment_type, 'description') else None
                    }
                }
            )
        except Exception as e:
            return Response(
                {
                    "message": "Failed to update segment type.",
                    "error": str(e)
                },
                status=status.HTTP_400_BAD_REQUEST
            )


class SegmentTypeDeleteView(APIView):
    """Delete a specific segment type
    
    Note: Cannot delete a segment type if it has associated segments.
    Delete all segments of this type first.
    """
    
    permission_classes = [IsAuthenticated]
    
    def delete(self, request, pk):
        try:
            segment_type = XX_SegmentType.objects.get(segment_id=pk)
        except XX_SegmentType.DoesNotExist:
            return Response(
                {"message": "Segment type not found."},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Check if segment type has associated segments
        segment_count = XX_Segment.objects.filter(segment_type=segment_type).count()
        
        if segment_count > 0:
            return Response(
                {
                    "message": f"Cannot delete segment type. It has {segment_count} associated segments.",
                    "suggestion": "Delete all segments of this type first, or reassign them to another segment type."
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        segment_type_name = segment_type.segment_name
        segment_type.delete()
        
        return Response(
            {
                "message": f"Segment type '{segment_type_name}' deleted successfully.",
            },
            status=status.HTTP_200_OK
        )

class SegmentTypeDeleteAllView(APIView):
    """Delete all segment types and their associated segments
    
    WARNING: This will delete ALL segment types and ALL segments in the system.
    Use with extreme caution.
    """

    permission_classes = [IsAuthenticated]

    def delete(self, request):
        from account_and_entitys.models import XX_BalanceReportSegment
        
        total_segment_types = XX_SegmentType.objects.count()
        total_balance_segments = XX_BalanceReportSegment.objects.count()

        # Delete in correct order to handle PROTECT foreign keys:
        # 1. Delete XX_BalanceReportSegment (has PROTECT FKs to SegmentType and Segment)
        # 2. Delete XX_Segment (has PROTECT FKs from BalanceReportSegment)
        # 3. Delete XX_SegmentType (has PROTECT FKs from BalanceReportSegment)
        XX_BalanceReportSegment.objects.all().delete()
        XX_Segment.objects.all().delete()
        XX_SegmentType.objects.all().delete()

        return Response(
            {
                "message": "All segment types, segments, and balance report segments have been deleted successfully.",
                "total_segment_types_deleted": total_segment_types,
                "total_balance_segments_deleted": total_balance_segments,
            },
            status=status.HTTP_200_OK
        )
# Segment are now stored in XX_Segment_XX table

##################################################################################### Fully dynamic segment view##############################################################

class SegmentListView(APIView):
    """Fully dynamic view to list segments based on configured segment types in XX_SegmentType table
    
    Works with any number of segment types (not limited to 3) and retrieves segment type info dynamically.
    
    Query Parameters:
    - segment_type: REQUIRED - Segment type ID to retrieve (from XX_SegmentType table)
                    Can also use segment_name (e.g., 'Entity', 'Account', 'Project', or any custom type)
    - search: Filter by code or alias (optional)
    - filter: Segment filter type (optional, default='all')
        * 'all': All segments regardless of level (default)
        * 'root' or 'zero_level': Only root segments (level=0)
        * 'leaf' or 'children': Only leaf segments (segments with no children)
        * 'exclude_leaf': All segments except leaf segments (only parents)
        * 'exclude_root': All segments except root segments (only non-root)
    """

    permission_classes = [IsAuthenticated]
    pagination_class = EntityPagination

    def get(self, request):
        search_query = request.query_params.get("search", None)
        filter_type = request.query_params.get("filter", "all").lower()
        segment_type_param = request.query_params.get("segment_type", None)

        # Validate segment_type parameter
        if not segment_type_param:
            # Get available segment types for error message
            available_types = XX_SegmentType.objects.all().values('segment_id', 'segment_name')
            types_list = [f"{t['segment_id']}={t['segment_name']}" for t in available_types]
            
            return Response(
                {
                    "message": "segment_type parameter is required.",
                    "available_segment_types": types_list,
                    "data": []
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # Try to find segment type by ID or by name
        segment_type_obj = None
        
        # First try as ID (integer)
        try:
            segment_type_id = int(segment_type_param)
            segment_type_obj = XX_SegmentType.objects.filter(segment_id=segment_type_id).first()
        except (ValueError, TypeError):
            # If not an integer, try as name
            segment_type_obj = XX_SegmentType.objects.filter(
                segment_name__iexact=segment_type_param
            ).first()
        
        # If still not found, return error with available types
        if not segment_type_obj:
            available_types = XX_SegmentType.objects.all().values('segment_id', 'segment_name')
            types_list = [f"{t['segment_id']}={t['segment_name']}" for t in available_types]
            
            return Response(
                {
                    "message": f"Invalid segment_type '{segment_type_param}'. Not found in configured segment types.",
                    "available_segment_types": types_list,
                    "data": []
                },
                status=status.HTTP_404_NOT_FOUND
            )

        # Start with base query for segments
        segments = XX_Segment.objects.filter(
            segment_type=segment_type_obj,
            is_active=True
        )

        # Apply filter based on filter_type parameter
        if filter_type in ['root', 'zero_level']:
            # Only root segments (level=0)
            segments = segments.filter(level=0)
        
        elif filter_type in ['leaf', 'children']:
            # Only leaf segments (segments that are not parents to any other segment)
            # Get all codes that are referenced as parent_code
            parent_codes = XX_Segment.objects.filter(
                segment_type=segment_type_obj,
                parent_code__isnull=False
            ).exclude(parent_code='').values_list('parent_code', flat=True).distinct()
            
            # Only include segments whose code does NOT appear as a parent
            segments = segments.exclude(code__in=parent_codes)
        
        elif filter_type == 'exclude_leaf':
            # Exclude leaf segments (only show segments that ARE parents to other segments)
            # Get all codes that are referenced as parent_code
            parent_codes = XX_Segment.objects.filter(
                segment_type=segment_type_obj,
                parent_code__isnull=False
            ).exclude(parent_code='').values_list('parent_code', flat=True).distinct()
            
            # Only include segments whose code DOES appear as a parent
            segments = segments.filter(code__in=parent_codes)
        
        elif filter_type == 'exclude_root':
            # Exclude root segments (only show non-root segments with level > 0)
            segments = segments.exclude(level=0)
        
        elif filter_type == 'all':
            # All segments - no additional filter needed (default)
            pass
        
        else:
            # Invalid filter type, default to all
            pass

        # Order by code
        segments = segments.order_by("code")

        # Apply search filter on both code and alias
        if search_query:
            segments = segments.filter(
                Q(code__icontains=search_query) | 
                Q(alias__icontains=search_query)
            )

        # Use the new SegmentValueListSerializer
        serializer = SegmentValueListSerializer(segments, many=True)

        return Response(
            {
                "message": f"{segment_type_obj.segment_name} retrieved successfully.", 
                "data": serializer.data,
                "segment_type": segment_type_obj.segment_name,
                "segment_type_id": segment_type_obj.segment_id,
                "filter_applied": filter_type,
                "total_count": segments.count()
            }
        )

class SegmentCreateView(APIView):
    """Unified dynamic view to create segments for any segment type
    
    Request Body:
    - segment_type: REQUIRED - Segment type ID or name (e.g., 1, 2, 3, "Account", "Project")
    - code: REQUIRED - Segment code
    - parent_code: OPTIONAL - Parent segment code (for hierarchy, auto-calculates level)
    - alias: OPTIONAL - Segment display name/alias
    - is_active: OPTIONAL - Active status (defaults to True)
    - envelope_amount: OPTIONAL - Budget/envelope amount for this segment (if applicable)
    
    Note: 
    - level is automatically calculated based on parent_code (0 for root, parent.level + 1 for children)
    - If you want to override level, include it in request body (not recommended)
    """

    permission_classes = [IsAuthenticated]

    def post(self, request):
        segment_type_param = request.data.get("segment_type")
        
        # Validate segment_type parameter
        if not segment_type_param:
            available_types = XX_SegmentType.objects.all().values('segment_id', 'segment_name')
            types_list = [f"{t['segment_id']}={t['segment_name']}" for t in available_types]
            
            return Response(
                {
                    "message": "segment_type is required in request body.",
                    "available_segment_types": types_list,
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Try to find segment type by ID or by name
        segment_type_obj = None
        
        # First try as ID (integer)
        try:
            segment_type_id = int(segment_type_param)
            segment_type_obj = XX_SegmentType.objects.filter(segment_id=segment_type_id).first()
        except (ValueError, TypeError):
            # If not an integer, try as name
            segment_type_obj = XX_SegmentType.objects.filter(
                segment_name__iexact=segment_type_param
            ).first()
        
        # If still not found, return error with available types
        if not segment_type_obj:
            available_types = XX_SegmentType.objects.all().values('segment_id', 'segment_name')
            types_list = [f"{t['segment_id']}={t['segment_name']}" for t in available_types]
            
            return Response(
                {
                    "message": f"Invalid segment_type '{segment_type_param}'. Not found in configured segment types.",
                    "available_segment_types": types_list,
                },
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Validate required fields
        code = request.data.get("code")
        if not code:
            return Response(
                {
                    "message": "code is required in request body.",
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if segment already exists
        existing_segment = XX_Segment.objects.filter(
            segment_type=segment_type_obj,
            code=code
        ).first()
        
        if existing_segment:
            return Response(
                {
                    "message": f"Segment with code '{code}' already exists for segment type '{segment_type_obj.segment_name}'.",
                    # "existing_segment": SegmentValueListSerializer(existing_segment).data
                },
                status=status.HTTP_409_CONFLICT
            )
        
        # Calculate level based on parent_code (or use provided level if user wants to override)
        parent_code = request.data.get("parent_code")
        level = request.data.get("level")  # Allow manual override if provided
        
        if parent_code:
            parent_segment = XX_Segment.objects.filter(
                segment_type=segment_type_obj,
                code=parent_code
            ).first()
            
            if parent_segment:
                # Auto-calculate level from parent (unless manually overridden)
                if level is None:
                    level = parent_segment.level + 1
            else:
                return Response(
                    {
                        "message": f"Parent segment with code '{parent_code}' not found for segment type '{segment_type_obj.segment_name}'.",
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
        else:
            # No parent means root level
            if level is None:
                level = 0
        
        # Prepare data for segment creation
        segment_data = {
            "segment_type": segment_type_obj,
            "code": code,
            "parent_code": parent_code,
            "alias": request.data.get("alias"),
            "level": level,
            "is_active": request.data.get("is_active", True),
        }
        
        # Handle envelope_amount if provided
        envelope_amount = request.data.get("envelope_amount")
        if envelope_amount is not None:
            try:
                from decimal import Decimal
                # Convert to Decimal for validation
                segment_data["envelope_amount"] = Decimal(str(envelope_amount))
            except (ValueError, TypeError) as e:
                return Response(
                    {
                        "message": "Invalid envelope_amount. Must be a valid number.",
                        "error": str(e)
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        # Create the segment
        try:
            segment = XX_Segment.objects.create(**segment_data)
            
            return Response(
                {
                    "message": f"{segment_type_obj.segment_name} created successfully.",
                    "data": SegmentValueListSerializer(segment).data,
                    "segment_type": segment_type_obj.segment_name,
                    "segment_type_id": segment_type_obj.segment_id,
                },
                status=status.HTTP_201_CREATED,
            )
        except Exception as e:
            return Response(
                {
                    "message": f"Failed to create {segment_type_obj.segment_name}.",
                    "error": str(e)
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

class SegmentDetailView(APIView):
    """Unified dynamic view to retrieve a specific segment by ID
    
    Works with any segment type. Returns full segment details including level and envelope_amount.
    """

    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            segment = XX_Segment.objects.get(pk=pk)
        except XX_Segment.DoesNotExist:
            return Response(
                {"message": "Segment not found."}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        serializer = SegmentValueListSerializer(segment)
        return Response(
            {
                "message": f"{segment.segment_type.segment_name} details retrieved successfully.",
                "data": serializer.data,
                "segment_type": segment.segment_type.segment_name,
                "segment_type_id": segment.segment_type.segment_id,
            }
        )

class SegmentUpdateView(APIView):
    """Unified dynamic view to update a specific segment
    
    Supports updating:
    - code (must remain unique within segment type)
    - parent_code (auto-recalculates level)
    - alias
    - envelope_amount
    - is_active
    - level (auto-calculated if parent_code changes, can be overridden)
    """

    permission_classes = [IsAuthenticated]

    def put(self, request, pk):
        try:
            segment = XX_Segment.objects.get(pk=pk)
        except XX_Segment.DoesNotExist:
            return Response(
                {"message": "Segment not found."}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Check if code is being changed and if it conflicts
        new_code = request.data.get("code")
        if new_code and new_code != segment.code:
            # Check for duplicate code in same segment type
            existing = XX_Segment.objects.filter(
                segment_type=segment.segment_type,
                code=new_code
            ).exclude(pk=pk).first()
            
            if existing:
                return Response(
                    {
                        "message": f"Segment with code '{new_code}' already exists for segment type '{segment.segment_type.segment_name}'.",
                    },
                    status=status.HTTP_409_CONFLICT
                )
        
        # Handle parent_code change and level recalculation
        new_parent_code = request.data.get("parent_code")
        new_level = request.data.get("level")
        
        # If parent_code is being changed, recalculate level
        if new_parent_code is not None and new_parent_code != segment.parent_code:
            if new_parent_code:
                # Find the new parent
                parent_segment = XX_Segment.objects.filter(
                    segment_type=segment.segment_type,
                    code=new_parent_code
                ).first()
                
                if parent_segment:
                    # Auto-calculate level unless manually overridden
                    if new_level is None:
                        request.data['level'] = parent_segment.level + 1
                else:
                    return Response(
                        {
                            "message": f"Parent segment with code '{new_parent_code}' not found for segment type '{segment.segment_type.segment_name}'.",
                        },
                        status=status.HTTP_400_BAD_REQUEST
                    )
            else:
                # Parent removed, becomes root
                if new_level is None:
                    request.data['level'] = 0
        
        # Handle envelope_amount validation
        envelope_amount = request.data.get("envelope_amount")
        if envelope_amount is not None:
            try:
                from decimal import Decimal
                request.data['envelope_amount'] = Decimal(str(envelope_amount))
            except (ValueError, TypeError) as e:
                return Response(
                    {
                        "message": "Invalid envelope_amount. Must be a valid number.",
                        "error": str(e)
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        # Update segment
        for field in ['code', 'parent_code', 'alias', 'level', 'envelope_amount', 'is_active']:
            if field in request.data:
                setattr(segment, field, request.data[field])
        
        try:
            segment.save()
            
            return Response(
                {
                    "message": f"{segment.segment_type.segment_name} updated successfully.",
                    "data": SegmentValueListSerializer(segment).data,
                    "segment_type": segment.segment_type.segment_name,
                    "segment_type_id": segment.segment_type.segment_id,
                }
            )
        except Exception as e:
            return Response(
                {
                    "message": f"Failed to update {segment.segment_type.segment_name}.",
                    "error": str(e)
                },
                status=status.HTTP_400_BAD_REQUEST
            )

class SegmentDeleteView(APIView):
    """Unified dynamic view to delete a specific segment
    
    Deletes a segment by ID. Works with any segment type.
    Note: Be careful deleting parent segments as it may orphan children.
    """

    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        try:
            segment = XX_Segment.objects.get(pk=pk)
        except XX_Segment.DoesNotExist:
            return Response(
                {"message": "Segment not found."}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Check if segment has children
        has_children = XX_Segment.objects.filter(
            segment_type=segment.segment_type,
            parent_code=segment.code
        ).exists()
        
        if has_children:
            return Response(
                {
                    "message": f"Cannot delete segment. It has child segments that depend on it.",
                    "suggestion": "Delete child segments first, or remove their parent_code references."
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        segment_type_name = segment.segment_type.segment_name
        segment_code = segment.code
        
        segment.delete()
        
        return Response(
            {
                "message": f"{segment_type_name} '{segment_code}' deleted successfully.",
            }, 
            status=status.HTTP_200_OK
        )

class SegmentDeleteAllView(APIView):
    """Delete all segment types and their associated segments
    
    WARNING: This will delete ALL segment types and ALL segments in the system.
    Use with extreme caution.
    """

    permission_classes = [IsAuthenticated]

    def delete(self, request):
        from account_and_entitys.models import XX_BalanceReportSegment
        
        total_segment = XX_Segment.objects.count()
        total_balance_segments = XX_BalanceReportSegment.objects.count()

        # Delete in correct order to handle PROTECT foreign keys:
        # 1. Delete XX_BalanceReportSegment (has PROTECT FK to Segment)
        # 2. Delete XX_Segment
        XX_BalanceReportSegment.objects.all().delete()
        XX_Segment.objects.all().delete()

        return Response(
            {
                "message": "All segments and balance report segments have been deleted successfully.",
                "total_segments_deleted": total_segment,
                "total_balance_segments_deleted": total_balance_segments,
            },
            status=status.HTTP_200_OK
        )
    
class SegmentBulkUploadView(APIView):
    
    """Unified dynamic bulk upload for any segment type via Excel file
    
    Expects:
    - Query parameter: segment_type (ID or name)
    - Excel file with columns: Code | ParentCode | Alias | EnvelopeAmount (optional)
    
    Features:
    - Works with any segment type (not limited to 3)
    - Auto-calculates hierarchy levels
    - Supports optional envelope_amount column
    - Upserts segments (creates new, updates existing)
    - Validates parent existence
    - Returns detailed summary
    """

    permission_classes = [IsAuthenticated]

    def post(self, request):
        """Handle file upload and process segments dynamically
        
        Excel file format:
        Row 1 (optional header): Code | ParentCode | Alias | EnvelopeAmount
        Row 2+: Actual data rows
        
        Note: EnvelopeAmount column is optional. If present, must be numeric or empty.
        """
        uploaded_file = request.FILES.get("file")
        segment_type_param = request.query_params.get("segment_type")

        if not uploaded_file:
            return Response(
                {"message": "No file uploaded."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not segment_type_param:
            available_types = XX_SegmentType.objects.all().values('segment_id', 'segment_name')
            types_list = [f"{t['segment_id']}={t['segment_name']}" for t in available_types]
            
            return Response(
                {
                    "message": "segment_type query parameter is required.",
                    "available_segment_types": types_list,
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # Find segment type by ID or name
        segment_type_obj = None
        try:
            segment_type_id = int(segment_type_param)
            segment_type_obj = XX_SegmentType.objects.filter(segment_id=segment_type_id).first()
        except (ValueError, TypeError):
            segment_type_obj = XX_SegmentType.objects.filter(
                segment_name__iexact=segment_type_param
            ).first()
        
        if not segment_type_obj:
            available_types = XX_SegmentType.objects.all().values('segment_id', 'segment_name')
            types_list = [f"{t['segment_id']}={t['segment_name']}" for t in available_types]
            
            return Response(
                {
                    "message": f"Invalid segment_type '{segment_type_param}'. Not found in configured segment types.",
                    "available_segment_types": types_list,
                },
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            from openpyxl import load_workbook
            from django.db import transaction
            from decimal import Decimal, InvalidOperation

            wb = load_workbook(filename=uploaded_file, read_only=True, data_only=True)
            sheet = wb.active

            created = 0
            updated = 0
            skipped = 0
            errors = []

            with transaction.atomic():
                first = True
                for row in sheet.iter_rows(values_only=True):
                    # Skip entirely empty rows
                    if not row or all(
                        [
                            c is None or (isinstance(c, str) and c.strip() == "")
                            for c in row
                        ]
                    ):
                        continue

                    # Normalize columns: code, parent_code, alias, envelope_amount (optional)
                    code = str(row[0]).strip() if row[0] is not None else None
                    parent_code = (
                        str(row[1]).strip()
                        if len(row) > 1 and row[1] is not None and str(row[1]).strip()
                        else None
                    )
                    alias = (
                        str(row[2]).strip()
                        if len(row) > 2 and row[2] is not None and str(row[2]).strip()
                        else None
                    )
                    envelope_amount = (
                        row[3]
                        if len(row) > 3 and row[3] is not None
                        else None
                    )

                    # Detect and skip header row
                    if first:
                        first = False
                        header_like = False
                        # Treat as header if first cell looks non-numeric or contains typical header words
                        if code and (
                            not any(ch.isdigit() for ch in code) or
                            code.lower() in ['code', 'segment', 'id', 'key', 'value']
                        ):
                            header_like = True
                        if header_like:
                            continue

                    # Validate code is present
                    if not code:
                        skipped += 1
                        continue

                    # Calculate level based on parent_code
                    level = 0
                    if parent_code:
                        parent_segment = XX_Segment.objects.filter(
                            segment_type=segment_type_obj,
                            code=parent_code
                        ).first()
                        
                        if parent_segment:
                            level = parent_segment.level + 1
                        else:
                            # Parent doesn't exist yet - will be skipped or error
                            errors.append({
                                "code": code,
                                "error": f"Parent code '{parent_code}' not found. Process parent rows first."
                            })
                            skipped += 1
                            continue

                    # Parse envelope_amount if provided
                    envelope_value = None
                    if envelope_amount is not None:
                        try:
                            # Handle various formats: string, number, etc.
                            envelope_str = str(envelope_amount).strip()
                            if envelope_str:
                                envelope_value = Decimal(envelope_str)
                        except (InvalidOperation, ValueError) as e:
                            errors.append({
                                "code": code,
                                "error": f"Invalid envelope_amount '{envelope_amount}': {str(e)}"
                            })
                            skipped += 1
                            continue

                    # Upsert: update if exists, else create
                    try:
                        obj, created_flag = XX_Segment.objects.update_or_create(
                            segment_type=segment_type_obj,
                            code=code,
                            defaults={
                                "parent_code": parent_code,
                                "alias": alias,
                                "level": level,
                                "envelope_amount": envelope_value,
                                "is_active": True,
                            },
                        )
                        if created_flag:
                            created += 1
                        else:
                            updated += 1
                    except Exception as row_err:
                        errors.append(
                            {"code": code, "error": str(row_err)}
                        )
                        skipped += 1

            summary = {
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "errors": errors,
                "segment_type": segment_type_obj.segment_name,
                "segment_type_id": segment_type_obj.segment_id,
            }

            return Response(
                {
                    "status": "ok", 
                    "message": f"Processed {created + updated} {segment_type_obj.segment_name} segments successfully.",
                    "summary": summary
                }, 
                status=status.HTTP_200_OK
            )

        except Exception as e:
            return Response(
                {
                    "status": "error", 
                    "message": str(e),
                    "segment_type": segment_type_obj.segment_name if segment_type_obj else None,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        
class Download_segment_values_from_oracle(APIView):
    """Download segment values from Oracle and update local XX_Segment table dynamically
    
    Query Parameters:
    - segment_type: REQUIRED - Segment type ID or name (e.g., 1, 2, 3, "Account", "Project")
    
    Features:
    - Works with any segment type (not limited to 3)
    - Fetches segments from Oracle via stored procedure
    - Upserts segments into local XX_Segment table
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
       
        oracle_maanger=OracleBalanceReportManager()

        oracle_maanger.download_segment_values_and_load_to_database(1)
        
    

        return Response("done")


# PivotFund views
class PivotFundListView(APIView):
    """List all pivot funds"""

    permission_classes = [IsAuthenticated]
    pagination_class = EntityPagination

    def get(self, request):
        # Allow filtering by entity, account, and year
        entity_id = request.query_params.get("entity")
        account_id = request.query_params.get("account")
        project_id = request.query_params.get("project")
        year = request.query_params.get("year")

        pivot_funds = XX_PivotFund.objects.all()

        if entity_id:
            pivot_funds = pivot_funds.filter(entity=entity_id)
        if account_id:
            pivot_funds = pivot_funds.filter(account=account_id)
        if project_id:
            pivot_funds = pivot_funds.filter(project=project_id)
        if year:
            pivot_funds = pivot_funds.filter(year=year)

        # Order by year, entity, account
        pivot_funds = pivot_funds.order_by(
            "-year", "entity__entity", "account__account", "project__project"
        )

        # Handle pagination
        paginator = self.pagination_class()
        paginated_funds = paginator.paginate_queryset(pivot_funds, request)
        serializer = PivotFundSerializer(paginated_funds, many=True)

        return paginator.get_paginated_response(serializer.data)

class PivotFundCreateView(APIView):
    """Create a new pivot fund"""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        # Handle batch creation
        if isinstance(request.data, list):
            created_funds = []
            errors = []

            for index, fund_data in enumerate(request.data):
                serializer = PivotFundSerializer(data=fund_data)
                if serializer.is_valid():
                    fund = serializer.save()
                    created_funds.append(PivotFundSerializer(fund).data)
                else:
                    errors.append(
                        {"index": index, "errors": serializer.errors, "data": fund_data}
                    )

            response_data = {
                "message": f"Created {len(created_funds)} pivot funds, with {len(errors)} errors.",
                "created": created_funds,
                "errors": errors,
            }

            if errors and not created_funds:
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
            elif errors:
                return Response(response_data, status=status.HTTP_207_MULTI_STATUS)
            else:
                return Response(response_data, status=status.HTTP_201_CREATED)

        # Handle single creation
        else:
            serializer = PivotFundSerializer(data=request.data)
            if serializer.is_valid():
                fund = serializer.save()
                return Response(
                    {
                        "message": "Pivot fund created successfully.",
                        "data": PivotFundSerializer(fund).data,
                    },
                    status=status.HTTP_201_CREATED,
                )
            return Response(
                {
                    "message": "Failed to create pivot fund.",
                    "errors": serializer.errors,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )


class PivotFundDetailView(APIView):
    """Retrieve a specific pivot fund"""

    permission_classes = [IsAuthenticated]

    def get_object(self, entity, account, project):
        try:

            return XX_PivotFund.objects.get(
                entity=entity, account=account, project=project
            )

        except XX_PivotFund.DoesNotExist:

            return None

    def get(self, request):

        entity = request.query_params.get("entity_id")
        account = request.query_params.get("account_id")
        project = request.query_params.get("project_id")
        print(entity, account, project)
        pivot_fund = self.get_object(entity, account, project)

        if pivot_fund is None:
            return Response(
                {"message": "Pivot fund not found."}, status=status.HTTP_200_OK
            )
        serializer = PivotFundSerializer(pivot_fund)
        return Response(
            {
                "message": "Pivot fund details retrieved successfully.",
                "data": serializer.data,
            }
        )


class PivotFundUpdateView(APIView):
    """Update a specific pivot fund"""

    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return XX_PivotFund.objects.get(pk=pk)
        except XX_PivotFund.DoesNotExist:
            return None

    def put(self, request, pk):
        pivot_fund = self.get_object(pk)
        if pivot_fund is None:
            return Response(
                {"message": "Pivot fund not found."}, status=status.HTTP_404_NOT_FOUND
            )
        serializer = PivotFundSerializer(pivot_fund, data=request.data)
        if serializer.is_valid():
            updated_fund = serializer.save()
            return Response(
                {
                    "message": "Pivot fund updated successfully.",
                    "data": PivotFundSerializer(updated_fund).data,
                }
            )
        return Response(
            {"message": "Failed to update pivot fund.", "errors": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST,
        )


class PivotFundDeleteView(APIView):
    """Delete a specific pivot fund"""

    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return XX_PivotFund.objects.get(pk=pk)
        except XX_PivotFund.DoesNotExist:
            return None

    def delete(self, request, pk):
        pivot_fund = self.get_object(pk)
        if pivot_fund is None:
            return Response(
                {"message": "Pivot fund not found."}, status=status.HTTP_404_NOT_FOUND
            )
        pivot_fund.delete()
        return Response(
            {"message": "Pivot fund deleted successfully."}, status=status.HTTP_200_OK
        )


# ADJD Transaction Audit views


class TransactionAuditListView(APIView):
    """List all ADJD transaction audit records"""

    permission_classes = [IsAuthenticated]
    pagination_class = EntityPagination

    def get(self, request):
        audit_records = XX_TransactionAudit.objects.all().order_by("-id")

        # Handle pagination
        paginator = self.pagination_class()
        paginated_records = paginator.paginate_queryset(audit_records, request)
        serializer = TransactionAuditSerializer(paginated_records, many=True)

        return paginator.get_paginated_response(serializer.data)


class TransactionAuditCreateView(APIView):
    """Create a new ADJD transaction audit record"""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = TransactionAuditSerializer(data=request.data)
        if serializer.is_valid():
            audit_record = serializer.save()
            return Response(
                {
                    "message": "Audit record created successfully.",
                    "data": TransactionAuditSerializer(audit_record).data,
                },
                status=status.HTTP_201_CREATED,
            )
        return Response(
            {"message": "Failed to create audit record.", "errors": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST,
        )


class TransactionAuditDetailView(APIView):
    """Retrieve a specific ADJD transaction audit record"""

    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return XX_TransactionAudit.objects.get(pk=pk)
        except XX_TransactionAudit.DoesNotExist:
            return None

    def get(self, request, pk):
        audit_record = self.get_object(pk)
        if audit_record is None:
            return Response(
                {"message": "Audit record not found."}, status=status.HTTP_404_NOT_FOUND
            )
        serializer = TransactionAuditSerializer(audit_record)
        return Response(
            {
                "message": "Audit record details retrieved successfully.",
                "data": serializer.data,
            }
        )


class TransactionAuditUpdateView(APIView):
    """Update a specific ADJD transaction audit record"""

    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return XX_TransactionAudit.objects.get(pk=pk)
        except XX_TransactionAudit.DoesNotExist:
            return None

    def put(self, request, pk):
        audit_record = self.get_object(pk)
        if audit_record is None:
            return Response(
                {"message": "Audit record not found."}, status=status.HTTP_404_NOT_FOUND
            )
        serializer = TransactionAuditSerializer(audit_record, data=request.data)
        if serializer.is_valid():
            updated_record = serializer.save()
            return Response(
                {
                    "message": "Audit record updated successfully.",
                    "data": TransactionAuditSerializer(updated_record).data,
                }
            )
        return Response(
            {"message": "Failed to update audit record.", "errors": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST,
        )


class TransactionAuditDeleteView(APIView):
    """Delete a specific ADJD transaction audit record"""

    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return XX_TransactionAudit.objects.get(pk=pk)
        except XX_TransactionAudit.DoesNotExist:
            return None

    def delete(self, request, pk):
        audit_record = self.get_object(pk)
        if audit_record is None:
            return Response(
                {"message": "Audit record not found."}, status=status.HTTP_404_NOT_FOUND
            )
        audit_record.delete()
        return Response(
            {"message": "Audit record deleted successfully."}, status=status.HTTP_200_OK
        )




class list_ACCOUNT_ENTITY_LIMIT(APIView):
    """List all ADJD transaction audit records"""

    permission_classes = [IsAuthenticated]
    pagination_class = EntityPagination

    def get(self, request):
        # Change "enity_id" to "entity_id"
        entity_id = request.query_params.get("cost_center")
        account_id = request.query_params.get("account_id")
        project_id = request.query_params.get("project_id")

        audit_records = XX_ACCOUNT_ENTITY_LIMIT.objects.filter(
            entity_id=entity_id
        ).order_by("-id")
        audit_records = audit_records.annotate(
            account_id_str=Cast("account_id", CharField())
        )
        audit_records = audit_records.annotate(
            project_id_str=Cast("project_id", CharField())
        )

        if account_id:
            audit_records = audit_records.filter(
                account_id_str__icontains=str(account_id)
            )
        if project_id:
            audit_records = audit_records.filter(
                project_id_str__icontains=str(project_id)
            )

        # Handle pagination
        paginator = self.pagination_class()
        paginated_records = paginator.paginate_queryset(audit_records, request)
        serializer = AccountEntityLimitSerializer(paginated_records, many=True)

        data = [
            {
                "id": record["id"],
                "account": record["account_id"],
                "project": record["project_id"],
                "is_transer_allowed_for_source": record[
                    "is_transer_allowed_for_source"
                ],
                "is_transer_allowed_for_target": record[
                    "is_transer_allowed_for_target"
                ],
                "is_transer_allowed": record["is_transer_allowed"],
                "source_count": record["source_count"],
                "target_count": record["target_count"],
            }
            for record in serializer.data
        ]

        return paginator.get_paginated_response(data)


class AccountEntityLimitAPI(APIView):
    """Handle both listing and creation of ACCOUNT_ENTITY_LIMIT records"""

    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]  # For file upload support

    def get(self, request):
        """List all records with optional filtering by cost_center"""
        entity_id = request.query_params.get("cost_center")

        audit_records = XX_ACCOUNT_ENTITY_LIMIT.objects.filter(
            entity_id=entity_id
        ).order_by("-id")

        paginator = self.pagination_class()
        paginated_records = paginator.paginate_queryset(audit_records, request)
        serializer = AccountEntityLimitSerializer(paginated_records, many=True)

        data = [
            {
                "id": record["id"],
                "account": record["account_id"],
                "project": record["project_id"],
                "is_transfer_allowed_for_source": record[
                    "is_transfer_allowed_for_source"
                ],
                "is_transfer_allowed_for_target": record[
                    "is_transfer_allowed_for_target"
                ],
                "is_transfer_allowed": record["is_transfer_allowed"],
                "source_count": record["source_count"],
                "target_count": record["target_count"],
            }
            for record in serializer.data
        ]

        return paginator.get_paginated_response(data)

    def post(self, request):
        """Handle both single record creation and bulk upload via file"""
        # Check if file is present for bulk upload
        uploaded_file = request.FILES.get("file")

        if uploaded_file:
            return self._handle_file_upload(uploaded_file)
        else:
            return self._handle_single_record(request.data)

    def _handle_file_upload(self, file):
        """Process Excel file for bulk creation"""
        try:
            # Read Excel file
            df = pd.read_excel(file)

            # Clean column names (convert to lowercase and strip whitespace)
            df.columns = df.columns.str.strip().str.lower()
            df = df.replace([np.nan, pd.NA, pd.NaT, "", "NULL", "null"], None)

            # Convert to list of dictionaries
            records = df.to_dict("records")

            created_count = 0
            errors = []

            with transaction.atomic():
                for idx, record in enumerate(records, start=1):
                    try:
                        serializer = AccountEntityLimitSerializer(data=record)
                        if serializer.is_valid():
                            serializer.save()
                            created_count += 1
                        else:
                            errors.append(
                                {
                                    "row": idx,
                                    "errors": serializer.errors,
                                    "data": record,
                                }
                            )
                    except Exception as e:
                        errors.append({"row": idx, "error": str(e), "data": record})

            response = {
                "status": "success",
                "created_count": created_count,
                "error_count": len(errors),
                "errors": errors if errors else None,
            }

            return Response(
                response,
                status=(
                    status.HTTP_201_CREATED
                    if created_count
                    else status.HTTP_400_BAD_REQUEST
                ),
            )

        except Exception as e:
            return Response(
                {"status": "error", "message": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )

    def _handle_single_record(self, data):
        """Handle single record creation"""
        serializer = AccountEntityLimitSerializer(data=data)

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class UpdateAccountEntityLimit(APIView):
    """Update a specific account entity limit."""

    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return XX_ACCOUNT_ENTITY_LIMIT.objects.get(pk=pk)
        except XX_ACCOUNT_ENTITY_LIMIT.DoesNotExist:
            return None

    def put(self, request):

        pk = request.query_params.get("pk")
        limit_record = self.get_object(pk)
        if limit_record is None:
            return Response(
                {"message": "Limit record not found."}, status=status.HTTP_404_NOT_FOUND
            )
        serializer = AccountEntityLimitSerializer(limit_record, data=request.data)
        if serializer.is_valid():
            updated_record = serializer.save()
            return Response(
                {
                    "message": "Limit record updated successfully.",
                    "data": AccountEntityLimitSerializer(updated_record).data,
                }
            )
        return Response(
            {"message": "Failed to update limit record.", "errors": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST,
        )


class DeleteAccountEntityLimit(APIView):
    """Delete a specific account entity limit."""

    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return XX_ACCOUNT_ENTITY_LIMIT.objects.get(pk=pk)
        except XX_ACCOUNT_ENTITY_LIMIT.DoesNotExist:
            return None

    def delete(self, request, pk):
        limit_record = self.get_object(pk)
        if limit_record is None:
            return Response(
                {"message": "Limit record not found."}, status=status.HTTP_404_NOT_FOUND
            )
        limit_record.delete()
        return Response(
            {"message": "Limit record deleted successfully."}, status=status.HTTP_200_OK
        )


# MainCurrency views


class RefreshBalanceReportView(APIView):
    """API view to refresh balance report data from Oracle"""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        """Trigger balance report refresh"""
        from .utils import refresh_balance_report_data

        budget_name = request.data.get("control_budget_name", "MIC_HQ_MONTHLY")
        period_name = request.data.get("Period_name", "sep-25")

        try:
            print("Starting balance report refresh...")
            print(f"Budget: {budget_name}, Period: {period_name}")
            result = refresh_balance_report_data(budget_name, period_name)
            if result["success"]:
                return Response(
                    {
                        "success": True,
                        "message": result["message"],
                        "data": {
                            "created_count": result["details"].get("created_count", 0),
                            "deleted_count": result["details"].get("deleted_count", 0),
                            "error_count": result["details"].get("error_count", 0),
                            "budget_name": budget_name,
                        },
                    },
                    status=status.HTTP_200_OK,
                )
            else:
                return Response(
                    {
                        "success": False,
                        "message": result["message"],
                        "errors": result.get("details", {}).get("errors", []),
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": f"Error refreshing balance report: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def get(self, request):
        """Get balance report refresh status"""
        from .models import XX_BalanceReport

        try:
            total_records = XX_BalanceReport.objects.count()
            latest_record = XX_BalanceReport.objects.order_by("-created_at").first()
            periods = list(
                XX_BalanceReport.objects.values_list(
                    "as_of_period", flat=True
                ).distinct()
            )

            return Response(
                {
                    "success": True,
                    "data": {
                        "total_records": total_records,
                        "available_periods": periods,
                        "latest_update": (
                            latest_record.created_at if latest_record else None
                        ),
                        "last_period": (
                            latest_record.as_of_period if latest_record else None
                        ),
                    },
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": f"Error getting balance report status: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class BalanceReportListView(APIView):
    """List balance report data with filtering"""

    permission_classes = [IsAuthenticated]
    pagination_class = EntityPagination

    def get(self, request):
        """Get balance report data with optional filtering"""
        from .models import XX_BalanceReport
        from .serializers import BalanceReportSerializer
        from .utils import extract_unique_segments_from_data

        try:
            control_budget_name = request.query_params.get("control_budget_name")
            period_name = request.query_params.get("as_of_period")

            # Check if user wants only unique segments
            extract_segments = (
                request.query_params.get("extract_segments", "").lower() == "true"
            )

            # Get data from Oracle
            data = get_oracle_report_data(control_budget_name, period_name)

            # Validate Oracle response structure
            if not isinstance(data, dict) or not data.get("success"):
                message = (
                    data.get("message")
                    if isinstance(data, dict)
                    else "Unexpected response"
                )
                return Response(
                    {
                        "success": False,
                        "message": f"Failed to retrieve report data: {message}",
                        "data": [],
                    },
                    status=status.HTTP_502_BAD_GATEWAY,
                )

            records = data.get("data", [])

            if extract_segments:
                unique_segments = extract_unique_segments_from_data(data)

                # Enrich segments with names (aliases); fallback to code when missing
                cost_centers = unique_segments.get("Cost_Center", []) or []
                accounts = unique_segments.get("Account", []) or []
                projects = unique_segments.get("Project", []) or []

                # Fetch aliases in bulk
                entity_alias_map = {
                    str(e.entity): (e.alias_default or str(e.entity))
                    for e in XX_Entity.objects.filter(entity__in=cost_centers)
                }
                account_alias_map = {
                    str(a.account): (a.alias_default or str(a.account))
                    for a in XX_Account.objects.filter(account__in=accounts)
                }
                project_alias_map = {
                    str(p.project): (p.alias_default or str(p.project))
                    for p in XX_Project.objects.filter(project__in=projects)
                }

                enriched = {
                    "Cost_Center": [
                        {
                            "code": str(code),
                            "name": entity_alias_map.get(str(code), str(code)),
                        }
                        for code in cost_centers
                    ],
                    "Account": [
                        {
                            "code": str(code),
                            "name": account_alias_map.get(str(code), str(code)),
                        }
                        for code in accounts
                    ],
                    "Project": [
                        {
                            "code": str(code),
                            "name": project_alias_map.get(str(code), str(code)),
                        }
                        for code in projects
                    ],
                    "total_records": unique_segments.get("total_records", 0),
                    "unique_combinations": unique_segments.get(
                        "unique_combinations", 0
                    ),
                }
                return Response(
                    {
                        "success": True,
                        "message": "Unique segments extracted successfully",
                        "data": enriched,
                    },
                    status=status.HTTP_200_OK,
                )

            # Otherwise, return the full data with unique segments included
            unique_segments = extract_unique_segments_from_data(data)

            # Enrich segments with names (aliases); fallback to code when missing
            cost_centers = unique_segments.get("Cost_Center", []) or []
            accounts = unique_segments.get("Account", []) or []
            projects = unique_segments.get("Project", []) or []

            entity_alias_map = {
                str(e.entity): (e.alias_default or str(e.entity))
                for e in XX_Entity.objects.filter(entity__in=cost_centers)
            }
            account_alias_map = {
                str(a.account): (a.alias_default or str(a.account))
                for a in XX_Account.objects.filter(account__in=accounts)
            }
            project_alias_map = {
                str(p.project): (p.alias_default or str(p.project))
                for p in XX_Project.objects.filter(project__in=projects)
            }

            enriched_unique_segments = {
                "Cost_Center": [
                    {
                        "code": str(code),
                        "name": entity_alias_map.get(str(code), str(code)),
                    }
                    for code in cost_centers
                ],
                "Account": [
                    {
                        "code": str(code),
                        "name": account_alias_map.get(str(code), str(code)),
                    }
                    for code in accounts
                ],
                "Project": [
                    {
                        "code": str(code),
                        "name": project_alias_map.get(str(code), str(code)),
                    }
                    for code in projects
                ],
                "total_records": unique_segments.get("total_records", 0),
                "unique_combinations": unique_segments.get("unique_combinations", 0),
            }

            return Response(
                {
                    "success": True,
                    "message": "Balance report data retrieved successfully",
                    "data": {
                        "records": records,
                        "unique_segments": enriched_unique_segments,
                        "total_records": len(records),
                    },
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": f"Error retrieving balance report data: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class BalanceReportSegmentsView(APIView):
    """API to get all unique segments from balance report"""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Get unique values for segment1, segment2, and segment3"""
        from .models import XX_BalanceReport

        try:
            # Get unique segments with filters
            segment1_filter = request.query_params.get("segment1")
            segment2_filter = request.query_params.get("segment2")

            queryset = XX_BalanceReport.objects.all()

            # Apply filters if provided
            if segment1_filter:
                queryset = queryset.filter(segment1=segment1_filter)
            if segment2_filter:
                queryset = queryset.filter(segment2=segment2_filter)

            # Get unique values for each segment
            segment1_values = list(
                XX_BalanceReport.objects.filter(segment1__isnull=False)
                .values_list("segment1", flat=True)
                .distinct()
                .order_by("segment1")
            )

            segment2_values = list(
                queryset.filter(segment2__isnull=False)
                .values_list("segment2", flat=True)
                .distinct()
                .order_by("segment2")
            )

            segment3_values = list(
                queryset.filter(segment3__isnull=False)
                .values_list("segment3", flat=True)
                .distinct()
                .order_by("segment3")
            )

            return Response(
                {
                    "success": True,
                    "data": {
                        "segment1": segment1_values,
                        "segment2": segment2_values,
                        "segment3": segment3_values,
                        "total_combinations": queryset.count(),
                    },
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {"success": False, "message": f"Error retrieving segments: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class BalanceReportOracleSegmentsView(APIView):
    """API to get unique segments from Oracle balance report data"""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Get unique segments from Oracle data with optional filtering"""
        from .utils import extract_unique_segments_from_data

        try:
            # Get filter parameters
            control_budget_name = request.query_params.get("control_budget_name")
            period_name = request.query_params.get("as_of_period")
            segment1_filter = request.query_params.get("segment1")
            segment2_filter = request.query_params.get("segment2")
            segment3_filter = request.query_params.get("segment3")

            # Get data from Oracle with filters
            oracle_data = get_oracle_report_data(
                control_budget_name,
                period_name,
                segment1_filter,
                segment2_filter,
                segment3_filter,
            )

            if not oracle_data:
                return Response(
                    {
                        "success": False,
                        "message": "No data found with the specified filters",
                    },
                    status=status.HTTP_404_NOT_FOUND,
                )

            # Extract unique segments from the Oracle data
            unique_segments = extract_unique_segments_from_data(oracle_data)

            return Response(
                {
                    "success": True,
                    "message": f"Successfully extracted unique segments from {len(oracle_data)} records",
                    "data": unique_segments,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": f"Error extracting segments from Oracle data: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def post(self, request):
        """Extract unique segments from provided balance report data"""
        from .utils import extract_unique_segments_from_data

        try:
            # Get data from request body
            balance_data = request.data.get("data", [])

            if not balance_data:
                return Response(
                    {
                        "success": False,
                        "message": "No balance report data provided in request body",
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Extract unique segments from the provided data
            unique_segments = extract_unique_segments_from_data(balance_data)

            return Response(
                {
                    "success": True,
                    "message": f"Successfully extracted unique segments from {len(balance_data)} records",
                    "data": unique_segments,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": f"Error extracting segments from provided data: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class BalanceReportFinancialDataView(APIView):
    """API to get financial data for specific segment combination"""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Get financial data for a specific segment1, segment2, segment3 combination"""
        from .models import XX_BalanceReport
        from django.db.models import Sum, Avg, Count

        try:
            from .utils import refresh_balance_report_data

            budget_name = request.data.get("control_budget_name", "MIC_HQ_MONTHLY")
            period_name = request.data.get("Period_name", "sep-25")

            try:
                result = refresh_balance_report_data(
                    budget_name, period_name=period_name
                )
            except Exception as e:
                return Response(
                    {
                        "success": False,
                        "message": f"Error refreshing balance report data: {str(e)}",
                    },
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

            # Get segments from query parameters
            segment1 = request.query_params.get("segment1")
            segment2 = request.query_params.get("segment2")
            segment3 = request.query_params.get("segment3")

            # Validate that all segments are provided
            if not all([segment1, segment2, segment3]):
                return Response(
                    {
                        "success": False,
                        "message": "All three segments (segment1, segment2, segment3) are required",
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Filter records by the segments
            queryset = XX_BalanceReport.objects.filter(
                segment1=segment1, segment2=segment2, segment3=segment3
            )

            if not queryset.exists():
                return Response(
                    {
                        "success": False,
                        "message": f"No data found for segments: {segment1}/{segment2}/{segment3}",
                    },
                    status=status.HTTP_404_NOT_FOUND,
                )

            # Get the most recent record for this combination
            latest_record = queryset.order_by("-created_at").first()

            # Calculate aggregated data if multiple records exist
            aggregated_data = queryset.aggregate(
                total_actual_ytd=Sum("actual_ytd"),
                total_encumbrance_ytd=Sum("encumbrance_ytd"),
                total_funds_available=Sum("funds_available_asof"),
                total_other_ytd=Sum("other_ytd"),
                total_budget_ytd=Sum("budget_ytd"),
                avg_actual_ytd=Avg("actual_ytd"),
                avg_encumbrance_ytd=Avg("encumbrance_ytd"),
                avg_funds_available=Avg("funds_available_asof"),
                record_count=Count("id"),
            )

            # Prepare response data
            financial_data = {
                "segments": {
                    "segment1": segment1,
                    "segment2": segment2,
                    "segment3": segment3,
                },
                "latest_record": {
                    "control_budget_name": latest_record.control_budget_name,
                    "ledger_name": latest_record.ledger_name,
                    "as_of_period": latest_record.as_of_period,
                    "actual_ytd": (
                        float(latest_record.actual_ytd)
                        if latest_record.actual_ytd
                        else 0
                    ),
                    "encumbrance_ytd": (
                        float(latest_record.encumbrance_ytd)
                        if latest_record.encumbrance_ytd
                        else 0
                    ),
                    "funds_available_asof": (
                        float(latest_record.funds_available_asof)
                        if latest_record.funds_available_asof
                        else 0
                    ),
                    "other_ytd": (
                        float(latest_record.other_ytd) if latest_record.other_ytd else 0
                    ),
                    "budget_ytd": (
                        float(latest_record.budget_ytd)
                        if latest_record.budget_ytd
                        else 0
                    ),
                    "last_updated": latest_record.created_at,
                },
                "aggregated_totals": {
                    "total_actual_ytd": float(aggregated_data["total_actual_ytd"] or 0),
                    "total_encumbrance_ytd": float(
                        aggregated_data["total_encumbrance_ytd"] or 0
                    ),
                    "total_funds_available": float(
                        aggregated_data["total_funds_available"] or 0
                    ),
                    "total_other_ytd": float(aggregated_data["total_other_ytd"] or 0),
                    "total_budget_ytd": float(aggregated_data["total_budget_ytd"] or 0),
                    "record_count": aggregated_data["record_count"],
                },
                "calculated_metrics": {
                    "budget_utilization_percent": (
                        round(
                            (
                                float(aggregated_data["total_actual_ytd"] or 0)
                                / float(aggregated_data["total_budget_ytd"] or 1)
                            )
                            * 100,
                            2,
                        )
                        if aggregated_data["total_budget_ytd"]
                        else 0
                    ),
                    "funds_remaining": float(
                        aggregated_data["total_funds_available"] or 0
                    ),
                    "total_committed": float(aggregated_data["total_actual_ytd"] or 0)
                    + float(aggregated_data["total_encumbrance_ytd"] or 0),
                },
            }

            return Response(
                {"success": True, "data": financial_data}, status=status.HTTP_200_OK
            )

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": f"Error retrieving financial data: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def post(self, request):
        """Get financial data for multiple segment combinations"""
        from .models import XX_BalanceReport

        try:
            segment_combinations = request.data.get("segments", [])

            if not segment_combinations:
                return Response(
                    {
                        "success": False,
                        "message": 'Please provide segment combinations in the format: [{"segment1": "10001", "segment2": "2205403", "segment3": "CTRLCE1"}]',
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            results = []

            for combo in segment_combinations:
                segment1 = combo.get("segment1")
                segment2 = combo.get("segment2")
                segment3 = combo.get("segment3")

                if not all([segment1, segment2, segment3]):
                    results.append(
                        {
                            "segments": combo,
                            "success": False,
                            "message": "Missing segment values",
                        }
                    )
                    continue

                # Get data for this combination
                record = (
                    XX_BalanceReport.objects.filter(
                        segment1=segment1, segment2=segment2, segment3=segment3
                    )
                    .order_by("-created_at")
                    .first()
                )

                if record:
                    results.append(
                        {
                            "segments": {
                                "segment1": segment1,
                                "segment2": segment2,
                                "segment3": segment3,
                            },
                            "success": True,
                            "data": {
                                "actual_ytd": (
                                    float(record.actual_ytd) if record.actual_ytd else 0
                                ),
                                "encumbrance_ytd": (
                                    float(record.encumbrance_ytd)
                                    if record.encumbrance_ytd
                                    else 0
                                ),
                                "funds_available_asof": (
                                    float(record.funds_available_asof)
                                    if record.funds_available_asof
                                    else 0
                                ),
                                "other_ytd": (
                                    float(record.other_ytd) if record.other_ytd else 0
                                ),
                                "budget_ytd": (
                                    float(record.budget_ytd) if record.budget_ytd else 0
                                ),
                                "as_of_period": record.as_of_period,
                                "last_updated": record.created_at,
                            },
                        }
                    )
                else:
                    results.append(
                        {
                            "segments": combo,
                            "success": False,
                            "message": "No data found for this segment combination",
                        }
                    )

            return Response(
                {
                    "success": True,
                    "data": results,
                    "total_requested": len(segment_combinations),
                    "found": len([r for r in results if r["success"]]),
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": f"Error processing segment combinations: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class Single_BalanceReportView(APIView):
    """Retrieve a specific balance report record by ID"""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        control_budget_name = request.query_params.get("control_budget_name")
        period_name = request.query_params.get("as_of_period")
        segment1 = request.query_params.get("segment1")
        segment2 = request.query_params.get("segment2")
        segment3 = request.query_params.get("segment3")

        data = get_oracle_report_data(
            control_budget_name, period_name, segment1, segment2, segment3
        )

        if data is None:
            return Response(
                {"message": "Balance report record not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        return Response(
            {
                "message": "Balance report record details retrieved successfully.",
                "data": data,
            }
        )


# envlop
class Upload_ProjectEnvelopeView(APIView):
    """Upload project envelopes via Excel file"""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        """Handle file upload and process project envelopes

        Expects an Excel file where the first sheet has rows like:
        ProjectCodeWithAlias | EnvelopeNumber

        The view will upsert each row into `Project_Envelope`.
        """
        uploaded_file = request.FILES.get("file")

        if not uploaded_file:
            return Response(
                {"message": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from openpyxl import load_workbook
            from django.db import transaction
            from decimal import Decimal, InvalidOperation
            import math
            import re

            from .models import Project_Envelope

            wb = load_workbook(filename=uploaded_file, read_only=True, data_only=True)
            sheet = wb.active

            created = 0
            updated = 0
            skipped = 0
            errors = []

            def extract_project_code(raw_value):
                """Return the leading numeric project code from the provided cell."""
                if raw_value is None:
                    return None
                if isinstance(raw_value, int):
                    return str(raw_value)
                if isinstance(raw_value, float):
                    if math.isnan(raw_value):
                        return None
                    if raw_value.is_integer():
                        return str(int(raw_value))
                    return str(raw_value)
                value_str = str(raw_value).strip()
                match = re.match(r"^\s*(\d+)", value_str)
                return match.group(1) if match else None

            with transaction.atomic():
                for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if not row or all(
                        cell is None or (isinstance(cell, str) and cell.strip() == "")
                        for cell in row
                    ):
                        continue

                    project_code = extract_project_code(
                        row[0] if len(row) > 0 else None
                    )
                    envelope_raw = row[1] if len(row) > 1 else None

                    if not project_code:
                        skipped += 1
                        continue

                    if envelope_raw is None or (
                        isinstance(envelope_raw, str) and envelope_raw.strip() == ""
                    ):
                        skipped += 1
                        continue

                    try:
                        if isinstance(envelope_raw, str):
                            cleaned = re.sub(
                                r"[^0-9.\-]", "", envelope_raw.replace(",", "")
                            )
                            envelope_val = Decimal(cleaned) if cleaned != "" else None
                        else:
                            envelope_val = Decimal(envelope_raw)
                    except (InvalidOperation, TypeError):
                        errors.append(
                            {
                                "row": idx,
                                "project_code": project_code,
                                "error": f"Invalid envelope value: {envelope_raw}",
                            }
                        )
                        continue

                    if envelope_val is None:
                        skipped += 1
                        continue

                    try:
                        _, created_flag = Project_Envelope.objects.update_or_create(
                            project=project_code,
                            defaults={"envelope": envelope_val},
                        )
                        if created_flag:
                            created += 1
                        else:
                            updated += 1
                    except Exception as row_err:
                        errors.append(
                            {
                                "row": idx,
                                "project_code": project_code,
                                "error": str(row_err),
                            }
                        )

            summary = {
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "errors": errors,
            }
            return Response(
                {"status": "ok", "summary": summary}, status=status.HTTP_200_OK
            )

        except Exception as e:
            return Response(
                {"status": "error", "message": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )
class ProjectEnvelopeListView(APIView):
    """List all Projects with optional search"""

    permission_classes = [IsAuthenticated]
    pagination_class = EntityPagination

    def get(self, request):
        search_query = request.query_params.get("search", None)

        if not request.user.projects.exists():
            projects = XX_Project.objects.all().order_by("project")
            envelope_projects = []
            for proj in projects:
                if EnvelopeManager.Has_Envelope(proj.project):
                    envelope_projects.append(proj.project)
            projects = XX_Project.objects.filter(project__in=envelope_projects)
        else:
            projects = request.user.projects.all().order_by("project")
            all_projects = []
            all_projects.extend([proj.project for proj in projects])
            for proj in projects:
                all_projects.extend(
                    EnvelopeManager.get_all_children(
                        XX_Project.objects.all(), proj.project
                    )
                )
            envelope_projects = []
            for proj in all_projects:
                if EnvelopeManager.Has_Envelope(proj):
                    envelope_projects.append(proj)
            projects = XX_Project.objects.filter(project__in=envelope_projects)
        # projects = get_zero_level_projects(projects)

        if search_query:
            # Cast project (int) to string for filtering
            projects = projects.filter(
                Q(
                    project__icontains=search_query
                )  # works because Django auto casts to text in SQL
            )

        serializer = ProjectSerializer(projects, many=True)

        return Response(
            {"message": "Projects retrieved successfully.", "data": serializer.data}
        )
class UploadBudgetDataView(APIView):
    """Upload budget data via Excel file"""

    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request):
        """Handle file upload and upsert budget data rows."""
        uploaded_file = request.FILES.get("file")

        if not uploaded_file:
            return Response(
                {"message": "No file uploaded."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            from openpyxl import load_workbook
            from django.db import transaction
            from decimal import Decimal, InvalidOperation
            import math
            import re

            wb = load_workbook(filename=uploaded_file, read_only=True, data_only=True)
            sheet = wb.active

            created = 0
            updated = 0
            skipped = 0
            errors = []

            def extract_project_code(raw_value):
                """Return the project code from the provided cell."""
                if raw_value is None:
                    return None
                if isinstance(raw_value, int):
                    return str(raw_value)
                if isinstance(raw_value, float):
                    if math.isnan(raw_value):
                        return None
                    if raw_value.is_integer():
                        return str(int(raw_value))
                    return str(raw_value)
                return str(raw_value).strip() or None

            def normalize_account(raw_value):
                if raw_value is None:
                    return None
                if isinstance(raw_value, float):
                    if math.isnan(raw_value):
                        return None
                    if raw_value.is_integer():
                        return str(int(raw_value))
                    return str(raw_value)
                value_str = str(raw_value).strip()
                return value_str or None

            def parse_budget(raw_value, label):
                if raw_value is None:
                    return Decimal("0")
                if isinstance(raw_value, int):
                    return Decimal(raw_value)
                if isinstance(raw_value, float):
                    if math.isnan(raw_value):
                        return Decimal("0")
                    return Decimal(str(raw_value))
                value_str = str(raw_value).strip()
                if not value_str:
                    return Decimal("0")
                cleaned = re.sub(r"[^0-9.\-]", "", value_str)
                if cleaned in {"", "-", ".", "-.", ".-"}:
                    return Decimal("0")
                try:
                    return Decimal(cleaned)
                except InvalidOperation as exc:
                    raise ValueError(f"Invalid {label} value: {raw_value}") from exc

            with transaction.atomic():
                for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if not row or all(
                        cell is None or (isinstance(cell, str) and cell.strip() == "")
                        for cell in row
                    ):
                        continue

                    project_code = extract_project_code(
                        row[0] if len(row) > 0 else None
                    )
                    account_code = normalize_account(row[1] if len(row) > 1 else None)

                    if not project_code or not account_code:
                        skipped += 1
                        continue

                    try:
                        fy24_budget = parse_budget(
                            row[2] if len(row) > 2 else None,
                            "FY24 budget",
                        )
                        fy25_budget = parse_budget(
                            row[3] if len(row) > 3 else None,
                            "FY25 budget",
                        )
                    except ValueError as budget_err:
                        errors.append(
                            {
                                "row": idx,
                                "project": project_code,
                                "account": account_code,
                                "error": str(budget_err),
                            }
                        )
                        continue

                    try:
                        _, created_flag = Budget_data.objects.update_or_create(
                            project=project_code,
                            account=account_code,
                            defaults={
                                "FY24_budget": fy24_budget,
                                "FY25_budget": fy25_budget,
                            },
                        )
                        if created_flag:
                            created += 1
                        else:
                            updated += 1
                    except Exception as row_err:
                        errors.append(
                            {
                                "row": idx,
                                "project": project_code,
                                "account": account_code,
                                "error": str(row_err),
                            }
                        )

            summary = {
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "errors": errors,
            }
            return Response(
                {"status": "ok", "summary": summary},
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {"status": "error", "message": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )
class ActiveProjectsWithEnvelopeView(APIView):
    """Return active projects with their current envelope and totals."""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Query params: year, month, IsApproved
        # project_code = request.query_params.get("project_code", None)
        # if project_code is None:
        #     return Response(
        #         {"message": "project_code query parameter is required."},
        #         status=status.HTTP_400_BAD_REQUEST,
        #     )
        # project = XX_Project.objects.get(pk=project_code)
        # if project is None:
        #     return Response(
        #         {"message": "Project not found."}, status=status.HTTP_404_NOT_FOUND
        #     )
        # year = request.query_params.get("year", None)
        # month = request.query_params.get("month", None)
        if not request.user.projects.exists():
            project_code = "9000000"
        else:
            project_code = request.user.projects.first().project
        results = EnvelopeManager.Get_Current_Envelope_For_Project(
            project_code=project_code
        )

        if results and "project_totals" in results:
            # Transform into array of objects
            transformed_data = []

            # Extract data from the original structure
            for proj, data in results["project_totals"].items():
                transformed_data.append(
                    {
                        "project_code": proj,
                        "submitted_total": (
                            data["submitted"]["total"] if data["submitted"] else 0
                        ),
                        "approved_total": (
                            data["approved"]["total"] if data["approved"] else 0
                        ),
                    }
                )

            return Response(
                {
                    "message": "Active projects with envelope.",
                    "initial_envelope": results["initial_envelope"],
                    "current_envelope": results["current_envelope"],
                    "current_envelope_change_percentage": (
                        (results["current_envelope"] - results["initial_envelope"])
                        / results["initial_envelope"]
                        if results["initial_envelope"] != 0
                        else 0
                    ),
                    "estimated_envelope": results["estimated_envelope"],
                    "estimated_envelope_change_percentage": (
                        (results["estimated_envelope"] - results["initial_envelope"])
                        / results["initial_envelope"]
                        if results["initial_envelope"] != 0
                        else 0
                    ),
                    "data": transformed_data,
                }
            )
class ProjectWiseDashboardView(APIView):
    """Project-wise dashboard data"""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        entity_code = request.query_params.get("entity_code", None)
        if entity_code is None:
            return Response(
                {"message": "entity_code query parameter is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        results = EnvelopeManager.Get_Dashboard_Data_For_Entity(entity_code)
        return Response(
            {
                "message": "Project-wise dashboard data.",
                "data": results,
            }
        )
class AccountWiseDashboardView(APIView):
    """Account-wise dashboard data"""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        project_code = request.query_params.get("project_code", None)
        if project_code is None:
            return Response(
                {"message": "project_code query parameter is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        project = XX_Project.objects.get(pk=project_code)
        if project is None:
            return Response(
                {"message": "Project not found."}, status=status.HTTP_404_NOT_FOUND
            )
        results = EnvelopeManager.Get_Dashboard_Data_For_Project(project.project)
        return Response(
            {
                "message": "Project-wise dashboard data.",
                "data": results,
            }
        )
# Mapping
class UploadMappingExcelView(APIView):
    """
    Upload Excel file with Account and Entity mapping data.
    Expected sheets:
    - 'Account' sheet with columns: source_account, target_account
    - 'Entity' sheet with columns: source_entity, target_entity
    """

    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def get(self, request):

        data = get_mapping_for_fusion_data()
        return Response({"data": data})

    def post(self, request):
        try:
            # Check if file is provided
            if "file" not in request.FILES:
                return Response(
                    {"status": "error", "message": "No file provided"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            uploaded_file = request.FILES["file"]

            # Validate file extension
            if not uploaded_file.name.endswith((".xlsx", ".xls")):
                return Response(
                    {
                        "status": "error",
                        "message": "File must be Excel format (.xlsx or .xls)",
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Read Excel file
            excel_file = pd.ExcelFile(uploaded_file)
            sheet_names = excel_file.sheet_names

            results = {
                "account_mappings": {
                    "created": 0,
                    "updated": 0,
                    "skipped": 0,
                    "errors": [],
                },
                "entity_mappings": {
                    "created": 0,
                    "updated": 0,
                    "skipped": 0,
                    "errors": [],
                },
            }

            # Process Account sheet
            if "Account" in sheet_names:
                try:
                    account_df = pd.read_excel(uploaded_file, sheet_name="Account")

                    # Validate required columns
                    required_cols = ["Source", "Target"]
                    missing_cols = [
                        col for col in required_cols if col not in account_df.columns
                    ]

                    if missing_cols:
                        results["account_mappings"]["errors"].append(
                            f"Missing columns in Account sheet: {', '.join(missing_cols)}"
                        )
                    else:
                        # Process each row
                        with transaction.atomic():
                            for index, row in account_df.iterrows():
                                try:
                                    source_account = str(row["Source"]).strip()
                                    target_account = str(row["Target"]).strip()

                                    # Skip rows with empty values
                                    if (
                                        pd.isna(row["Source"])
                                        or pd.isna(row["Target"])
                                        or source_account == ""
                                        or target_account == ""
                                    ):
                                        results["account_mappings"]["skipped"] += 1
                                        continue

                                    # Create or update account mapping
                                    obj, created = (
                                        XX_ACCOUNT_mapping.objects.update_or_create(
                                            source_account=source_account,
                                            target_account=target_account,
                                            defaults={"is_active": True},
                                        )
                                    )

                                    if created:
                                        results["account_mappings"]["created"] += 1
                                    else:
                                        results["account_mappings"]["updated"] += 1

                                except Exception as row_err:
                                    results["account_mappings"]["errors"].append(
                                        f"Row {index + 2}: {str(row_err)}"
                                    )

                except Exception as sheet_err:
                    results["account_mappings"]["errors"].append(
                        f"Error processing Account sheet: {str(sheet_err)}"
                    )
            else:
                results["account_mappings"]["errors"].append(
                    "Account sheet not found in Excel file"
                )

            # Process Entity sheet
            if "Entity" in sheet_names:
                try:
                    entity_df = pd.read_excel(uploaded_file, sheet_name="Entity")

                    # Validate required columns
                    required_cols = ["Source", "Target"]
                    missing_cols = [
                        col for col in required_cols if col not in entity_df.columns
                    ]

                    if missing_cols:
                        results["entity_mappings"]["errors"].append(
                            f"Missing columns in Entity sheet: {', '.join(missing_cols)}"
                        )
                    else:
                        # Process each row
                        with transaction.atomic():
                            for index, row in entity_df.iterrows():
                                try:
                                    source_entity = str(row["Source"]).strip()
                                    target_entity = str(row["Target"]).strip()

                                    # Skip rows with empty values
                                    if (
                                        pd.isna(row["Source"])
                                        or pd.isna(row["Target"])
                                        or source_entity == ""
                                        or target_entity == ""
                                    ):
                                        results["entity_mappings"]["skipped"] += 1
                                        continue

                                    # Create or update entity mapping
                                    obj, created = (
                                        XX_Entity_mapping.objects.update_or_create(
                                            source_entity=source_entity,
                                            target_entity=target_entity,
                                            defaults={"is_active": True},
                                        )
                                    )

                                    if created:
                                        results["entity_mappings"]["created"] += 1
                                    else:
                                        results["entity_mappings"]["updated"] += 1

                                except Exception as row_err:
                                    results["entity_mappings"]["errors"].append(
                                        f"Row {index + 2}: {str(row_err)}"
                                    )

                except Exception as sheet_err:
                    results["entity_mappings"]["errors"].append(
                        f"Error processing Entity sheet: {str(sheet_err)}"
                    )
            else:
                results["entity_mappings"]["errors"].append(
                    "Entity sheet not found in Excel file"
                )

            # Determine overall status
            has_errors = (
                len(results["account_mappings"]["errors"]) > 0
                or len(results["entity_mappings"]["errors"]) > 0
            )

            response_status = (
                status.HTTP_207_MULTI_STATUS if has_errors else status.HTTP_200_OK
            )

            return Response(
                {
                    "status": "completed_with_errors" if has_errors else "success",
                    "message": (
                        "File processed successfully"
                        if not has_errors
                        else "File processed with some errors"
                    ),
                    "results": results,
                },
                status=response_status,
            )

        except Exception as e:
            return Response(
                {"status": "error", "message": f"Unexpected error: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

class UploadAccountMappingView(APIView):
    """Upload account mapping entries via Excel file"""

    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request):
        """Handle file upload and persist account mappings"""
        uploaded_file = request.FILES.get("file")

        if not uploaded_file:
            return Response(
                {"message": "No file uploaded."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            from openpyxl import load_workbook
            import math

            wb = load_workbook(filename=uploaded_file, read_only=True, data_only=True)
            sheet = wb.active

            created = 0
            duplicates = 0
            skipped = 0
            errors = []

            def normalize(value):
                if value is None:
                    return None
                if isinstance(value, float):
                    if math.isnan(value):
                        return None
                    if value.is_integer():
                        return str(int(value))
                return str(value).strip()

            with transaction.atomic():
                first_row = True
                for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if not row or all(
                        cell is None or (isinstance(cell, str) and cell.strip() == "")
                        for cell in row
                    ):
                        continue

                    source_account = normalize(row[0]) if len(row) > 0 else None
                    target_account = normalize(row[1]) if len(row) > 1 else None

                    if first_row:
                        first_row = False
                        lower_source = (source_account or "").lower()
                        lower_target = (target_account or "").lower()
                        if lower_source in {
                            "source",
                            "source_account",
                        } or lower_target in {"target", "target_account"}:
                            continue

                    if not source_account or not target_account:
                        skipped += 1
                        continue

                    try:
                        _, created_flag = Account_Mapping.objects.get_or_create(
                            source_account=source_account,
                            target_account=target_account,
                        )
                        if created_flag:
                            created += 1
                        else:
                            duplicates += 1
                    except Exception as row_err:
                        errors.append(
                            {
                                "row": idx,
                                "source_account": source_account,
                                "target_account": target_account,
                                "error": str(row_err),
                            }
                        )

            summary = {
                "created": created,
                "duplicates": duplicates,
                "skipped": skipped,
                "errors": errors,
            }

            return Response(
                {"status": "ok", "summary": summary},
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {"status": "error", "message": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )

class EntityMappingListView(APIView):
    """List all entity mappings"""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Implement logic to retrieve and return entity mappings
        data = get_mapping_for_fusion_data()
        return Response(
            {"message": "Entity mappings retrieved successfully.", "data": data}
        )


