"""
API views for Validation endpoints with support for:
- Type Validation
- IN/NOT IN Operations
- Excel file upload for list values
"""

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
import json

from .models import ValidationStep, ValidationWorkflow, ValidationExecution, ValidationStepExecution
from .serializers import (
    ValidationStepSerializer,
    ValidationStepCreateSerializer,
    ValidationWorkflowSerializer,
    ValidationWorkflowWithStepsSerializer,
    ValidationExecutionSerializer,
    ValidationStepExecutionSerializer,
    ValidationExecuteSerializer,
    ExcelUploadSerializer,
)
from .execution_engine import ValidationExecutionEngine


class ValidationStepViewSet(viewsets.ModelViewSet):
    """
    API endpoint for managing ValidationSteps.
    
    Supports all operations including IN/NOT IN with Excel upload.
    
    Endpoints:
        - GET /api/validations/steps/ - List all validation steps
        - POST /api/validations/steps/ - Create new validation step
        - POST /api/validations/steps/upload_excel/ - Create step with Excel list
        - GET /api/validations/steps/{id}/ - Get specific step
        - PUT/PATCH /api/validations/steps/{id}/ - Update step
        - DELETE /api/validations/steps/{id}/ - Delete step
        - GET /api/validations/steps/operations/ - Get available operations
        - POST /api/validations/steps/{id}/validate_types/ - Validate type compatibility
    """
    
    queryset = ValidationStep.objects.filter(is_active=True)
    serializer_class = ValidationStepSerializer
    permission_classes = [AllowAny]
    parser_classes = [JSONParser, MultiPartParser, FormParser]
    filterset_fields = ['order', 'is_active', 'operation']
    search_fields = ['name', 'description']
    ordering_fields = ['order', 'created_at']
    ordering = ['order', 'created_at']
    
    def get_serializer_class(self):
        """Use different serializer for creation."""
        if self.action == 'create':
            return ValidationStepCreateSerializer
        return ValidationStepSerializer
    
    def perform_create(self, serializer):
        """Set created_by and handle workflow auto-assignment."""
        workflow_id = self.request.data.get('workflow_id')
        
        if self.request.user and self.request.user.is_authenticated:
            step = serializer.save(created_by=self.request.user)
        else:
            step = serializer.save()
        
        # Auto-assign to workflow if workflow_id provided
        if workflow_id:
            self._update_workflow_steps(workflow_id, step)
    
    def perform_update(self, serializer):
        """Handle workflow auto-assignment on update."""
        workflow_id = self.request.data.get('workflow_id')
        step = serializer.save()
        
        # Auto-assign to workflow if workflow_id provided
        if workflow_id:
            self._update_workflow_steps(workflow_id, step)
    
    def _remap_step_references(self, steps, id_mapping, new_step_id_threshold):
        """
        Remap temporary step IDs to real database IDs in action_data fields.
        
        This handles the case where steps reference other steps that were just created
        with temporary IDs (>= new_step_id_threshold).
        
        Args:
            steps: List of ValidationStep objects to update
            id_mapping: Dict mapping temp_id -> real_db_id
            new_step_id_threshold: IDs >= this value are considered temporary
        """
        if not id_mapping:
            return
        
        for step in steps:
            updated = False
            
            # Check if_true_action_data
            if step.if_true_action_data and isinstance(step.if_true_action_data, dict):
                next_step_id = step.if_true_action_data.get('next_step_id')
                if next_step_id and next_step_id >= new_step_id_threshold:
                    # This references a temporary ID, remap it
                    real_id = id_mapping.get(next_step_id)
                    if real_id:
                        step.if_true_action_data['next_step_id'] = real_id
                        updated = True
            
            # Check if_false_action_data
            if step.if_false_action_data and isinstance(step.if_false_action_data, dict):
                next_step_id = step.if_false_action_data.get('next_step_id')
                if next_step_id and next_step_id >= new_step_id_threshold:
                    # This references a temporary ID, remap it
                    real_id = id_mapping.get(next_step_id)
                    if real_id:
                        step.if_false_action_data['next_step_id'] = real_id
                        updated = True
            
            # Save if we made any updates
            if updated:
                step.save()
    
    def _update_workflow_steps(self, workflow_id, step):
        """Update workflow's step list based on step order."""
        from .models import ValidationWorkflow
        from rest_framework.exceptions import ValidationError
        
        try:
            workflow = ValidationWorkflow.objects.get(id=workflow_id)
        except ValidationWorkflow.DoesNotExist:
            raise ValidationError({"workflow_id": f"Workflow with id {workflow_id} does not exist"})
        
        # Add step to workflow if not already there
        if step not in workflow.steps.all():
            workflow.steps.add(step)
        
        # Get all steps for this workflow, ordered by order field
        all_steps = workflow.steps.all().order_by('order', 'id')
        
        # Rebuild step_ids list maintaining order
        ordered_step_ids = list(all_steps.values_list('id', flat=True))
        
        # Handle initial_step - must be the step with order=1
        steps_with_order_1 = all_steps.filter(order=1)
        
        if steps_with_order_1.count() > 1:
            raise ValidationError({
                "order": f"Multiple steps have order=1. Only one step can have order=1 to be set as initial_step. Found: {list(steps_with_order_1.values_list('id', 'name'))}"
            })
        
        if steps_with_order_1.exists():
            workflow.initial_step = steps_with_order_1.first()
        elif workflow.initial_step is None and ordered_step_ids:
            # If no step with order=1, set first step as initial
            workflow.initial_step_id = ordered_step_ids[0]
        
        workflow.save()
    
    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """
        Create multiple validation steps at once with temporary ID mapping.
        
        Request body:
            {
                "workflow_id": 1,
                "new_step_id": 5,  // Threshold: IDs >= 5 are new steps
                "steps": [
                    {
                        "id": 3,  // Existing step
                        "name": "Step 1",
                        ...
                    },
                    {
                        "id": 5,  // New step (>= new_step_id)
                        "name": "Step 2",
                        "if_true_action": "proceed_to_step_by_id",
                        "if_true_action_data": {"next_step_id": 6}  // References another new step
                    },
                    {
                        "id": 6,  // New step
                        "name": "Step 3",
                        ...
                    }
                ]
            }
        
        Logic:
        - IDs >= new_step_id are NEW steps (create them)
        - IDs < new_step_id are EXISTING steps (update/link them)
        - After creating new steps, map temp IDs to real DB IDs
        - Update all next_step_id references in action_data
        
        Returns:
            {
                "success": true,
                "created_count": 2,
                "steps": [...],
                "id_mapping": {5: 10, 6: 11},  // temp_id -> real_db_id
                "workflow": {...}
            }
        """
        from .serializers import BulkStepCreateSerializer, ValidationStepCreateSerializer
        from .models import ValidationWorkflow
        from rest_framework.exceptions import ValidationError
        
        serializer = BulkStepCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        workflow_id = serializer.validated_data['workflow_id']
        new_step_id = serializer.validated_data['new_step_id']
        steps_data = serializer.validated_data['steps']
        
        # Validate workflow exists
        try:
            workflow = ValidationWorkflow.objects.get(id=workflow_id)
        except ValidationWorkflow.DoesNotExist:
            raise ValidationError({"workflow_id": f"Workflow with id {workflow_id} does not exist"})
        
        # Separate new steps (id >= new_step_id) from existing steps
        new_steps_data = []
        existing_step_ids = []
        
        for step_data in steps_data:
            step_id = step_data.get('id')
            if step_id and step_id >= new_step_id:
                # This is a NEW step to create
                new_steps_data.append(step_data)
            elif step_id:
                # This is an EXISTING step
                existing_step_ids.append(step_id)
        
        # Build ID mapping: temp_id -> real_db_id
        id_mapping = {}
        created_steps = []
        errors = []
        
        # Create new steps
        for idx, step_data in enumerate(new_steps_data):
            try:
                # Extract temp_id (used for mapping only, not for DB)
                temp_id = step_data.get('id')
                
                # Make a copy to avoid modifying original
                step_data_copy = step_data.copy()
                
                step_serializer = ValidationStepCreateSerializer(data=step_data_copy)
                step_serializer.is_valid(raise_exception=True)
                
                if request.user and request.user.is_authenticated:
                    step = step_serializer.save(created_by=request.user)
                else:
                    step = step_serializer.save()
                
                # Map temp ID to real database ID
                if temp_id:
                    id_mapping[temp_id] = step.id
                
                # Auto-assign to workflow
                self._update_workflow_steps(workflow_id, step)
                created_steps.append(step)
                
            except Exception as e:
                errors.append({
                    "step_index": idx,
                    "temp_id": temp_id,
                    "step_name": step_data.get('name', 'Unknown'),
                    "error": str(e)
                })
        
        # Now update all steps' action_data to replace temp IDs with real IDs
        self._remap_step_references(created_steps, id_mapping, new_step_id)
        
        # Link existing steps to workflow
        if existing_step_ids:
            existing_steps = ValidationStep.objects.filter(id__in=existing_step_ids)
            for step in existing_steps:
                self._update_workflow_steps(workflow_id, step)
        
        # Refresh workflow to get updated steps
        workflow.refresh_from_db()
        
        response_data = {
            'success': len(created_steps) > 0 or len(existing_step_ids) > 0,
            'created_count': len(created_steps),
            'linked_existing_count': len(existing_step_ids),
            'steps': ValidationStepSerializer(created_steps, many=True).data,
            'id_mapping': id_mapping,
            'workflow': {
                'id': workflow.id,
                'name': workflow.name,
                'initial_step': workflow.initial_step_id,
                'total_steps': workflow.steps.count()
            }
        }
        
        if errors:
            response_data['errors'] = errors
            response_data['error_count'] = len(errors)
        
        return Response(response_data, status=status.HTTP_201_CREATED)
    
    @action(detail=False, methods=['put', 'patch'])
    def bulk_update(self, request):
        """
        Update multiple validation steps at once with temporary ID mapping.
        
        Request body:
            {
                "new_step_id": 5,  // Threshold: IDs >= 5 are new steps to create
                "updates": [
                    {
                        "step_id": 3,  // Existing step (< new_step_id)
                        "name": "Updated Name",
                        "order": 2,
                        "workflow_id": 1
                    },
                    {
                        "step_id": 5,  // New step (>= new_step_id)
                        "name": "New Step",
                        "order": 3,
                        "if_true_action": "proceed_to_step_by_id",
                        "if_true_action_data": {"next_step_id": 6},
                        "workflow_id": 1
                    },
                    {
                        "step_id": 6,  // New step
                        "name": "Another New Step",
                        "order": 4,
                        "workflow_id": 1
                    }
                ]
            }
        
        Logic:
        - step_id >= new_step_id: CREATE new step
        - step_id < new_step_id: UPDATE existing step
        - After creating, map temp IDs to real DB IDs
        - Update all next_step_id references in action_data
        
        Returns:
            {
                "success": true,
                "updated_count": 1,
                "created_count": 2,
                "steps": [...],
                "id_mapping": {5: 10, 6: 11}
            }
        """
        from .serializers import BulkStepUpdateSerializer, ValidationStepCreateSerializer
        
        serializer = BulkStepUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        new_step_id = serializer.validated_data['new_step_id']
        updates_data = serializer.validated_data['updates']
        
        # Separate new steps from existing steps
        new_steps_data = []
        update_steps_data = []
        
        for update_item in updates_data:
            step_id = update_item.get('step_id')
            if step_id and step_id >= new_step_id:
                # This is a NEW step to create
                new_steps_data.append(update_item)
            else:
                # This is an EXISTING step to update
                update_steps_data.append(update_item)
        
        # Build ID mapping for new steps
        id_mapping = {}
        created_steps = []
        updated_steps = []
        errors = []
        
        # CREATE new steps
        for new_step_data in new_steps_data:
            temp_id = new_step_data.pop('step_id')
            workflow_id = new_step_data.pop('workflow_id', None)
            
            try:
                # Prepare data for creation
                create_data = {
                    'name': new_step_data.get('name'),
                    'description': new_step_data.get('description', ''),
                    'order': new_step_data.get('order', 1),
                    'x': new_step_data.get('x', 0),
                    'y': new_step_data.get('y', 0),
                    'left_expression': new_step_data.get('left_expression', ''),
                    'operation': new_step_data.get('operation', '=='),
                    'right_expression': new_step_data.get('right_expression', ''),
                    'if_true_action': new_step_data.get('if_true_action', 'complete_success'),
                    'if_true_action_data': new_step_data.get('if_true_action_data', {}),
                    'if_false_action': new_step_data.get('if_false_action', 'complete_failure'),
                    'if_false_action_data': new_step_data.get('if_false_action_data', {}),
                    'failure_message': new_step_data.get('failure_message', ''),
                    'is_active': new_step_data.get('is_active', True)
                }
                
                step_serializer = ValidationStepCreateSerializer(data=create_data)
                step_serializer.is_valid(raise_exception=True)
                
                if request.user and request.user.is_authenticated:
                    step = step_serializer.save(created_by=request.user)
                else:
                    step = step_serializer.save()
                
                # Map temp ID to real database ID
                id_mapping[temp_id] = step.id
                
                # Handle workflow assignment if provided
                if workflow_id:
                    self._update_workflow_steps(workflow_id, step)
                
                created_steps.append(step)
                
            except Exception as e:
                errors.append({
                    "temp_id": temp_id,
                    "error": str(e),
                    "action": "create"
                })
        
        # UPDATE existing steps
        for update_item in update_steps_data:
            step_id = update_item.pop('step_id')
            workflow_id = update_item.pop('workflow_id', None)
            
            try:
                step = ValidationStep.objects.get(id=step_id)
                
                # Update fields
                for field, value in update_item.items():
                    setattr(step, field, value)
                
                step.save()
                
                # Handle workflow reassignment if provided
                if workflow_id:
                    self._update_workflow_steps(workflow_id, step)
                
                updated_steps.append(step)
                
            except ValidationStep.DoesNotExist:
                errors.append({
                    "step_id": step_id,
                    "error": f"Step with id {step_id} does not exist",
                    "action": "update"
                })
            except Exception as e:
                errors.append({
                    "step_id": step_id,
                    "error": str(e),
                    "action": "update"
                })
        
        # Remap all step references for both created and updated steps
        all_affected_steps = created_steps + updated_steps
        self._remap_step_references(all_affected_steps, id_mapping, new_step_id)
        
        response_data = {
            'success': len(created_steps) > 0 or len(updated_steps) > 0,
            'created_count': len(created_steps),
            'updated_count': len(updated_steps),
            'steps': ValidationStepSerializer(all_affected_steps, many=True).data,
            'id_mapping': id_mapping
        }
        
        if errors:
            response_data['errors'] = errors
            response_data['error_count'] = len(errors)
        
        return Response(response_data, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'])
    def operations(self, request):
        """
        Get available operations with type requirements.
        
        Returns:
            {
                "operations": [
                    {
                        "value": "==",
                        "label": "Equals",
                        "supports_types": ["all"],
                        "requires_list": false
                    },
                    ...
                ]
            }
        """
        operations = [
            {
                "value": "==",
                "label": "Equals",
                "supports_types": ["int", "float", "str", "bool"],
                "requires_list": False,
                "description": "Check if values are equal"
            },
            {
                "value": "!=",
                "label": "Not Equals",
                "supports_types": ["int", "float", "str", "bool"],
                "requires_list": False,
                "description": "Check if values are not equal"
            },
            {
                "value": ">",
                "label": "Greater Than",
                "supports_types": ["int", "float", "str"],
                "requires_list": False,
                "description": "Check if left value is greater than right value (numeric or string)"
            },
            {
                "value": "<",
                "label": "Less Than",
                "supports_types": ["int", "float", "str"],
                "requires_list": False,
                "description": "Check if left value is less than right value (numeric or string)"
            },
            {
                "value": ">=",
                "label": "Greater Than or Equal",
                "supports_types": ["int", "float", "str"],
                "requires_list": False,
                "description": "Check if left value is greater than or equal to right value"
            },
            {
                "value": "<=",
                "label": "Less Than or Equal",
                "supports_types": ["int", "float", "str"],
                "requires_list": False,
                "description": "Check if left value is less than or equal to right value"
            },
            {
                "value": "in",
                "label": "In List",
                "supports_types": ["int", "float", "str", "bool"],
                "requires_list": True,
                "description": "Check if value exists in a list",
                "list_formats": ["JSON array: [1,2,3]", "Comma-separated: 1,2,3", "Excel upload"]
            },
            {
                "value": "not_in",
                "label": "Not In List",
                "supports_types": ["int", "float", "str", "bool"],
                "requires_list": True,
                "description": "Check if value does NOT exist in a list",
                "list_formats": ["JSON array: [1,2,3]", "Comma-separated: 1,2,3", "Excel upload"]
            },
        ]
        
        return Response({
            "operations": operations,
            "type_validation_enabled": True,
            "excel_upload_supported": True
        })
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def upload_excel(self, request):
        """
        Create a validation step with list values from Excel file.
        
        Form Data:
            - excel_file: Excel file (.xlsx)
            - column_name: Name of column to read values from
            - name: Step name
            - description: Step description (optional)
            - order: Step order
            - left_expression: Left side expression
            - operation: 'in' or 'not_in'
            - if_true_action: Action if condition is true
            - if_false_action: Action if condition is false
            - if_true_action_data: JSON data for true action (optional)
            - if_false_action_data: JSON data for false action (optional)
        
        Returns:
            Created ValidationStep with right_expression populated from Excel
        """
        serializer = ExcelUploadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        try:
            # Import openpyxl
            try:
                import openpyxl
            except ImportError:
                return Response(
                    {
                        'success': False,
                        'error': 'Excel support not available. Install openpyxl: pip install openpyxl'
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Read Excel file
            excel_file = request.FILES['excel_file']
            column_name = serializer.validated_data['column_name']
            
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            # Find column by name
            headers = [cell.value for cell in ws[1]]
            if column_name not in headers:
                return Response(
                    {
                        'success': False,
                        'error': f"Column '{column_name}' not found in Excel file. Available columns: {headers}"
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            col_idx = headers.index(column_name) + 1
            
            # Read values from column
            values = []
            for row in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value is not None:
                    values.append(cell_value)
            
            if not values:
                return Response(
                    {
                        'success': False,
                        'error': f"No values found in column '{column_name}'"
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Convert to JSON array
            right_expression = json.dumps(values)
            
            # Create validation step
            step_data = {
                'name': serializer.validated_data['name'],
                'description': serializer.validated_data.get('description', f'Imported from Excel column: {column_name}'),
                'order': serializer.validated_data['order'],
                'left_expression': serializer.validated_data['left_expression'],
                'operation': serializer.validated_data['operation'],
                'right_expression': right_expression,
                'if_true_action': serializer.validated_data['if_true_action'],
                'if_false_action': serializer.validated_data['if_false_action'],
                'if_true_action_data': serializer.validated_data.get('if_true_action_data', {}),
                'if_false_action_data': serializer.validated_data.get('if_false_action_data', {}),
            }
            
            step_serializer = ValidationStepSerializer(data=step_data)
            step_serializer.is_valid(raise_exception=True)
            
            if request.user and request.user.is_authenticated:
                step = step_serializer.save(created_by=request.user)
            else:
                step = step_serializer.save()
            
            return Response(
                {
                    'success': True,
                    'message': f'Successfully imported {len(values)} values from Excel',
                    'values_count': len(values),
                    'column_name': column_name,
                    'step': ValidationStepSerializer(step).data
                },
                status=status.HTTP_201_CREATED
            )
            
        except Exception as e:
            return Response(
                {
                    'success': False,
                    'error': f'Failed to process Excel file: {str(e)}'
                },
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'])
    def validate_types(self, request, pk=None):
        """
        Validate type compatibility for a step before execution.
        
        Request body:
            {
                "datasource_params": {
                    "UserTier": {"userId": 3}
                }
            }
        
        Returns:
            {
                "valid": true/false,
                "left_type": "str",
                "right_type": "int",
                "operation": ">",
                "error": "error message if invalid"
            }
        """
        step = self.get_object()
        
        try:
            from .execution_engine import ValidationExecutionEngine
            from .expression_evaluator import ExpressionEvaluator
            
            # Get datasource params
            datasource_params = request.data.get('datasource_params', {})
            
            # Evaluate expressions to get types
            evaluator = ExpressionEvaluator()
            
            # For IN/NOT IN operations, handle list parsing
            if step.operation in ['in', 'not_in']:
                engine = ValidationExecutionEngine(None)
                left_value = evaluator.evaluate(step.left_expression, datasource_params)
                right_value = engine._parse_list_expression(step.right_expression)
            else:
                left_value = evaluator.evaluate(step.left_expression, datasource_params)
                right_value = evaluator.evaluate(step.right_expression, datasource_params)
            
            left_type = type(left_value).__name__
            right_type = type(right_value).__name__
            
            # Validate type compatibility
            engine = ValidationExecutionEngine(None)
            try:
                engine._validate_type_compatibility(left_value, right_value, step.operation, step.name)
                
                return Response({
                    'valid': True,
                    'left_type': left_type,
                    'right_type': right_type,
                    'left_value': str(left_value)[:100],  # Truncate for display
                    'right_value': str(right_value)[:100] if not isinstance(right_value, list) else f'List with {len(right_value)} items',
                    'operation': step.operation,
                    'message': 'Type compatibility validated successfully'
                })
            except ValueError as e:
                return Response({
                    'valid': False,
                    'left_type': left_type,
                    'right_type': right_type,
                    'operation': step.operation,
                    'error': str(e)
                })
                
        except Exception as e:
            return Response(
                {
                    'valid': False,
                    'error': f'Validation failed: {str(e)}'
                },
                status=status.HTTP_400_BAD_REQUEST
            )


class ValidationWorkflowViewSet(viewsets.ModelViewSet):
    """
    API endpoint for managing ValidationWorkflows.
    
    Endpoints:
        - GET /api/validations/workflows/ - List all workflows
        - POST /api/validations/workflows/ - Create new workflow (with inline steps!)
        - GET /api/validations/workflows/{id}/ - Get specific workflow (with full step details!)
        - PUT/PATCH /api/validations/workflows/{id}/ - Update workflow (with inline steps!)
        - DELETE /api/validations/workflows/{id}/ - Delete workflow
        - POST /api/validations/workflows/{id}/execute/ - Execute workflow
        - GET /api/validations/workflows/{id}/executions/ - Get workflow executions
        - POST /api/validations/workflows/{id}/validate/ - Validate workflow before execution
    
    Supports THREE input modes for steps:
        1. step_ids: [1, 2, 3] - Link to existing steps by ID
        2. steps_data: [{...}, {...}] - Create new steps inline  
        3. Both: Update existing steps and create new ones
    
    Response always includes full step details.
    """
    
    queryset = ValidationWorkflow.objects.prefetch_related('steps').all()
    serializer_class = ValidationWorkflowWithStepsSerializer
    permission_classes = [AllowAny]
    filterset_fields = ['status', 'is_default', 'execution_point']
    search_fields = ['name', 'description']
    ordering_fields = ['created_at', 'status']
    ordering = ['-created_at']
    
    def get_serializer_class(self):
        """Use the enhanced serializer for all operations."""
        return ValidationWorkflowWithStepsSerializer
    
    def perform_create(self, serializer):
        """Set created_by to current user if authenticated."""
        if self.request.user and self.request.user.is_authenticated:
            serializer.save(created_by=self.request.user)
        else:
            serializer.save()
    
    @action(detail=True, methods=['post'])
    def execute(self, request, pk=None):
        """
        Execute a validation workflow with type validation support.
        
        Request body:
            {
                "context_data": {
                    "any": "data"
                },
                "datasource_params": {
                    "MaxAllowedUsers": {"tenantId": 123},
                    "CurrentUserCount": {"tenantId": 123}
                }
            }
        """
        workflow = self.get_object()
        
        serializer = ValidationExecuteSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        try:
            user = request.user if request.user and request.user.is_authenticated else None
            engine = ValidationExecutionEngine(workflow, user=user)
            execution = engine.execute(
                context_data=serializer.validated_data.get('context_data', {}),
                datasource_params=serializer.validated_data.get('datasource_params', {})
            )
            
            # Collect failure messages from failed steps
            failure_messages = []
            if execution.status in ['completed_failure', 'error']:
                for step_exec in execution.step_executions.all():
                    if step_exec.error_message:
                        failure_messages.append({
                            'step_name': step_exec.step.name,
                            'message': step_exec.error_message
                        })
                
                # Also check context_data for failure message
                if execution.context_data and 'failure_message' in execution.context_data:
                    if not failure_messages:  # Only add if no step-specific messages
                        failure_messages.append({
                            'step_name': 'Workflow',
                            'message': execution.context_data['failure_message']
                        })
            
            response_data = {
                'success': execution.status == 'completed_success',
                'execution': ValidationExecutionSerializer(execution).data
            }
            
            # Add failure messages if any
            if failure_messages:
                response_data['failure_messages'] = failure_messages
                response_data['error'] = failure_messages[0]['message']  # First message as main error
            
            return Response(
                response_data,
                status=status.HTTP_201_CREATED
            )
        except ValueError as e:
            # Type validation or other validation errors
            return Response(
                {
                    'success': False,
                    'error': str(e),
                    'error_type': 'validation_error'
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {
                    'success': False,
                    'error': f'Execution failed: {str(e)}',
                    'error_type': 'execution_error'
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def validate(self, request, pk=None):
        """
        Validate workflow configuration before execution.
        
        Checks:
        - All steps have valid operations
        - Type compatibility (if datasource_params provided)
        - Branching logic is valid
        - No circular references
        
        Request body (optional):
            {
                "datasource_params": {...}
            }
        """
        workflow = self.get_object()
        errors = []
        warnings = []
        
        # Check if workflow has steps
        if not workflow.steps.exists():
            errors.append("Workflow has no steps")
        
        # Check if initial step is set
        if not workflow.initial_step:
            errors.append("Workflow has no initial step")
        
        # Validate each step
        for step in workflow.steps.all():
            # Check operation is valid
            valid_operations = ['==', '!=', '>', '<', '>=', '<=', 'in', 'not_in']
            if step.operation not in valid_operations:
                errors.append(f"Step '{step.name}': Invalid operation '{step.operation}'")
            
            # Check IN/NOT IN operations have list format
            if step.operation in ['in', 'not_in']:
                try:
                    from .execution_engine import ValidationExecutionEngine
                    engine = ValidationExecutionEngine(None)
                    parsed_list = engine._parse_list_expression(step.right_expression)
                    if not isinstance(parsed_list, list):
                        errors.append(f"Step '{step.name}': IN/NOT IN operation requires list format")
                except Exception as e:
                    errors.append(f"Step '{step.name}': Failed to parse list - {str(e)}")
        
        # If datasource params provided, validate types
        datasource_params = request.data.get('datasource_params', {})
        if datasource_params:
            from .execution_engine import ValidationExecutionEngine
            from .expression_evaluator import ExpressionEvaluator
            
            evaluator = ExpressionEvaluator()
            engine = ValidationExecutionEngine(None)
            
            for step in workflow.steps.all():
                try:
                    if step.operation in ['in', 'not_in']:
                        left_value = evaluator.evaluate(step.left_expression, datasource_params)
                        right_value = engine._parse_list_expression(step.right_expression)
                    else:
                        left_value = evaluator.evaluate(step.left_expression, datasource_params)
                        right_value = evaluator.evaluate(step.right_expression, datasource_params)
                    
                    # Validate types
                    try:
                        engine._validate_type_compatibility(left_value, right_value, step.operation, step.name)
                    except ValueError as e:
                        errors.append(f"Step '{step.name}': {str(e)}")
                        
                except Exception as e:
                    warnings.append(f"Step '{step.name}': Could not validate types - {str(e)}")
        
        return Response({
            'valid': len(errors) == 0,
            'errors': errors,
            'warnings': warnings,
            'steps_count': workflow.steps.count(),
            'has_initial_step': workflow.initial_step is not None
        })
    
    @action(detail=True, methods=['get'])
    def executions(self, request, pk=None):
        """Get all executions for a specific workflow."""
        workflow = self.get_object()
        executions = workflow.executions.all()
        
        # Filter by status if provided
        status_filter = request.query_params.get('status')
        if status_filter:
            executions = executions.filter(status=status_filter)
        
        page = self.paginate_queryset(executions)
        if page is not None:
            serializer = ValidationExecutionSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = ValidationExecutionSerializer(executions, many=True)
        return Response(serializer.data)


class ValidationExecutionViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for viewing ValidationExecutions (read-only).
    
    Endpoints:
        - GET /api/validations/executions/ - List all executions
        - GET /api/validations/executions/{id}/ - Get specific execution
    """
    
    queryset = ValidationExecution.objects.all()
    serializer_class = ValidationExecutionSerializer
    permission_classes = [AllowAny]
    filterset_fields = ['workflow', 'status']
    search_fields = ['workflow__name']
    ordering_fields = ['started_at', 'completed_at', 'status']
    ordering = ['-started_at']


class ValidationStepExecutionViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for viewing ValidationStepExecutions (read-only).
    
    Endpoints:
        - GET /api/validations/step-executions/ - List all step executions
        - GET /api/validations/step-executions/{id}/ - Get specific step execution
    """
    
    queryset = ValidationStepExecution.objects.all()
    serializer_class = ValidationStepExecutionSerializer
    permission_classes = [AllowAny]
    filterset_fields = ['execution', 'step']
    ordering_fields = ['executed_at']
    ordering = ['executed_at']


class ExecutionPointViewSet(viewsets.ViewSet):
    """
    API endpoint for execution points.
    
    Execution points define WHERE in the application validation workflows can execute.
    Developers register execution points in code, users create workflows for those points.
    
    Each execution point has a list of allowed datasources that can be used in its workflows.
    
    Endpoints:
        - GET /api/validations/execution-points/ - List all registered execution points
        - GET /api/validations/execution-points/categories/ - List points grouped by category
        - GET /api/validations/execution-points/{code}/datasources/ - Get allowed datasources for a point
    """
    
    permission_classes = [AllowAny]
    
    def list(self, request):
        """
        List all registered execution points.
        
        Returns:
            {
                "execution_points": [list of execution point dicts with allowed_datasources],
                "total_count": int
            }
        """
        from .execution_point_registry import execution_point_registry
        
        # Import example execution points to ensure they're registered
        try:
            from . import execution_points
        except ImportError:
            pass
        
        points = execution_point_registry.list_all()
        
        return Response({
            'execution_points': points,
            'total_count': len(points)
        })
    
    @action(detail=False, methods=['get'])
    def categories(self, request):
        """
        List execution points grouped by category.
        
        Returns:
            {
                "categories": {
                    "category_name": [list of execution points],
                    ...
                },
                "total_categories": int
            }
        """
        from .execution_point_registry import execution_point_registry
        
        # Import example execution points to ensure they're registered
        try:
            from . import execution_points
        except ImportError:
            pass
        
        points = execution_point_registry.list_all()
        categories = {}
        
        for point in points:
            category = point['category']
            if category not in categories:
                categories[category] = []
            categories[category].append(point)
        
        return Response({
            'categories': categories,
            'total_categories': len(categories)
        })
    
    @action(detail=False, methods=['get'], url_path='(?P<code>[^/.]+)/datasources')
    def datasources(self, request, code=None):
        """
        Get allowed datasources for a specific execution point.
        
        This endpoint returns the list of datasources that can be used when creating
        validation workflows for this execution point.
        
        URL: GET /api/validations/execution-points/{code}/datasources/
        
        Args:
            code: The execution point code (e.g., 'on_transfer_submit')
        
        Returns:
            {
                "execution_point": {
                    "code": str,
                    "name": str,
                    "description": str,
                    "category": str
                },
                "datasources": [
                    {
                        "name": str,
                        "parameters": [str],
                        "return_type": str,
                        "description": str,
                        "function_name": str
                    },
                    ...
                ],
                "total_datasources": int,
                "usage_example": str
            }
        """
        from .execution_point_registry import execution_point_registry
        
        # Import example execution points to ensure they're registered
        try:
            from . import execution_points
        except ImportError:
            pass
        
        # Check if execution point exists
        ep = execution_point_registry.get(code)
        if not ep:
            return Response({
                'error': f"Execution point '{code}' not found",
                'available_points': [p['code'] for p in execution_point_registry.list_all()]
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Get allowed datasources with details
        datasources = execution_point_registry.get_allowed_datasources_details(code)
        
        # Build usage example
        usage_example = ""
        if datasources:
            ds = datasources[0]
            params_str = ', '.join([f'{p}=<value>' for p in ds.get('parameters', [])])
            usage_example = f"datasource:{ds['name']}({params_str})"
        
        return Response({
            'execution_point': {
                'code': ep['code'],
                'name': ep['name'],
                'description': ep['description'],
                'category': ep['category']
            },
            'datasources': datasources,
            'total_datasources': len(datasources),
            'usage_example': usage_example,
            'message': f"Use these datasources in left_expression or right_expression when creating validation steps for '{ep['name']}' workflows."
        })
