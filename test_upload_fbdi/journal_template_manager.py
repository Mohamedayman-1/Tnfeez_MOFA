import pandas as pd
import openpyxl
from openpyxl import load_workbook
import shutil
from pathlib import Path
import time
from typing import Dict, List, Any
from test_upload_fbdi.zip_fbdi import excel_to_csv_and_zip
import sys
from pathlib import Path

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from account_and_entitys.oracle import OracleSegmentMapper


def create_clean_journal_template(template_path: str, output_path: str = None) -> str:
    """
    Create a clean copy of the JournalImportTemplate with empty GL_INTERFACE data rows.

    Args:
        template_path: Path to the original JournalImportTemplate.xlsm
        output_path: Path for the clean template (optional, auto-generates if not provided)

    Returns:
        Path to the clean template file
    """
    template_file = Path(template_path)

    if not template_file.exists():
        raise FileNotFoundError(f"Template file not found: {template_path}")

    # Generate output path if not provided
    if output_path is None:
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        output_path = (
            template_file.parent / f"JournalImportTemplate_Clean_{timestamp}.xlsm"
        )
    else:
        output_path = Path(output_path)

    # Create a copy of the template
    shutil.copy2(template_file, output_path)
    print(f"Created clean template copy: {output_path}")

    # Open the copy and clear GL_INTERFACE data
    wb = load_workbook(output_path)

    if "GL_INTERFACE" in wb.sheetnames:
        gl_sheet = wb["GL_INTERFACE"]

        # Keep rows 1-4 (instructions and headers), delete data rows from row 5 onwards
        if gl_sheet.max_row > 4:
            gl_sheet.delete_rows(5, gl_sheet.max_row - 4)
            print(
                f"Cleared data rows from GL_INTERFACE sheet, keeping header structure"
            )

        # Save the changes
        wb.save(output_path)
        wb.close()

    return str(output_path)


def fill_journal_template_with_data(
    template_path: str,
    journal_data: List[Dict[str, Any]],
    output_path: str = None,
    auto_zip: bool = False,
) -> str:
    """
    Fill a clean journal template with data and optionally create ZIP.

    Args:
        template_path: Path to the clean JournalImportTemplate.xlsm
        journal_data: List of dictionaries containing journal entry data
        output_path: Path for the filled template (optional)
        auto_zip: Whether to automatically create ZIP file

    Returns:
        Path to the filled template file (or ZIP file if auto_zip=True)
    """
    template_file = Path(template_path)

    if not template_file.exists():
        raise FileNotFoundError(f"Template file not found: {template_path}")

    if not journal_data:
        raise ValueError("No journal data provided")

    # Generate output path if not provided
    if output_path is None:
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        output_path = template_file.parent / f"JournalImport_Filled_{timestamp}.xlsm"
    else:
        output_path = Path(output_path)

    # Create a copy for filling
    shutil.copy2(template_file, output_path)

    # Open the workbook and get the GL_INTERFACE sheet
    wb = load_workbook(output_path)
    gl_sheet = wb["GL_INTERFACE"]

    # Get the headers from row 4
    headers = []
    for col_num in range(1, gl_sheet.max_column + 1):
        header_cell = gl_sheet.cell(row=4, column=col_num)
        if header_cell.value is not None and str(header_cell.value).strip():
            clean_header = str(header_cell.value).lstrip("*").strip()
            headers.append(clean_header)
        else:
            break

    print(f"Found {len(headers)} headers in template")
    print(f"Filling template with {len(journal_data)} journal entries")

    # Fill data starting from row 5
    for row_idx, entry in enumerate(journal_data, start=5):
        for col_idx, header in enumerate(headers, start=1):
            # Get value from entry data
            value = entry.get(header, "")

            # Set the cell value
            cell = gl_sheet.cell(row=row_idx, column=col_idx)

            # Handle different data types appropriately
            if value is not None:
                cell.value = value
            else:
                cell.value = ""  # Empty for None
            # elif isinstance(value, str) and value.strip():
            #     cell.value = value.strip()
            # else:
            #     cell.value = ""  # Empty for None or empty strings

    # Save the filled template
    wb.save(output_path)
    wb.close()

    print(f"Template filled with data: {output_path}")

    # Optionally create ZIP file
    if auto_zip:
        zip_path = str(output_path).replace(".xlsm", ".zip")
        try:
            zip_result = excel_to_csv_and_zip(str(output_path), zip_path)
            print(f"ZIP file created: {zip_result}")
            return zip_result
        except Exception as e:
            print(f"Warning: Failed to create ZIP file: {e}")
            return str(output_path)

    return str(output_path)


def create_journal_from_scratch(
    template_path: str,
    journal_data: List[Dict[str, Any]],
    output_name: str = None,
    auto_zip: bool = True,
) -> str:
    """
    Complete workflow: Create clean template, fill with data, and optionally ZIP.

    Args:
        template_path: Path to the original JournalImportTemplate.xlsm
        journal_data: List of dictionaries containing journal entry data
        output_name: Base name for output files (optional)
        auto_zip: Whether to create ZIP file automatically

    Returns:
        Path to the final file (Excel or ZIP)
    """
    timestamp = time.strftime("%Y%m%d_%H%M%S")

    if output_name is None:
        output_name = f"JournalImport_{timestamp}"

    try:
        # Step 1: Create clean template
        clean_template = create_clean_journal_template(template_path)

        # Step 2: Fill with data
        filled_template_path = Path(template_path).parent / f"{output_name}.xlsm"
        result_path = fill_journal_template_with_data(
            clean_template, journal_data, str(filled_template_path), auto_zip=auto_zip
        )

        # Clean up temporary clean template
        Path(clean_template).unlink(missing_ok=True)

        return result_path

    except Exception as e:
        print(f"Error in journal creation workflow: {e}")
        raise


# Example usage function
def create_sample_journal_data(
    transfers, transaction_id=0, type="submit", group_id=0
) -> List[Dict[str, Any]]:
    """
    Create journal entry data with dynamic segment support.

    Args:
        transfers: List of XX_TransactionTransfer objects with segment relationships
        transaction_id: Transaction identifier
        type: "submit" or "reject" - determines debit/credit direction
        group_id: Interface group identifier for Oracle

    Returns:
        List of journal entries with dynamic SEGMENT1-SEGMENT30 columns

    Notes:
        - Uses OracleSegmentMapper to dynamically build segment columns
        - Supports 2-30 segments based on transaction configuration
        - Generates separate journal lines for each transfer (debit) and one offsetting line (credit)
    """
    # Initialize mapper
    mapper = OracleSegmentMapper()
    
    # Generate unique values that will be the same for all transfers in this function run
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    batch_name = f"BATCH_TRANSFER_{timestamp} transaction id={transaction_id}"
    batch_description = (
        f"Balance Transfer Batch created on {time.strftime('%Y-%m-%d %H:%M:%S')}"
    )
    journal_name = f"JOURNAL_TRANSFER_{timestamp} transaction id={transaction_id}"
    journal_description = f"Journal Entry for Balance Transfer - Created {time.strftime('%Y-%m-%d %H:%M:%S')}"

    sample_data = []
    total_debit = 0
    
    # Calculate total for offsetting entry
    for transfer in transfers:
        total_debit += (
            getattr(transfer, "from_center", 0) or 0
        )

    # Create journal lines for each transfer (FROM side - debit entries)
    for transfer in transfers:
        from_amount = getattr(transfer, "from_center", 0) or 0
        if from_amount > 0:
            # Base journal entry data (non-segment fields)
            journal_entry = {
                "Status Code": "NEW",
                "Ledger ID": "300000205309206",
                "Effective Date of Transaction": "2025-09-26",
                "Journal Source": "Allocations",
                "Journal Category": "Adjustment",
                "Currency Code": "AED",
                "Journal Entry Creation Date": "2025-09-26",
                "Actual Flag": "E",
                "Entered Debit Amount": (
                    from_amount if type == "submit" else ""
                ),
                "Entered Credit Amount": (
                    from_amount if type == "reject" else ""
                ),
                "REFERENCE1 (Batch Name)": batch_name,
                "REFERENCE2 (Batch Description)": batch_description,
                "REFERENCE4 (Journal Entry Name)": journal_name,
                "REFERENCE5 (Journal Entry Description)": journal_description,
                "REFERENCE10 (Journal Entry Line Description)": f"Debit transcation {transfer.transaction_id} transfer line {transfer.transfer_id}",
                "Encumbrance Type ID": "100000243328511",
                "Interface Group Identifier": group_id,
            }
            
            # Add dynamic segment columns using mapper
            # This builds SEGMENT1, SEGMENT2, ... SEGMENT30 based on transaction configuration
            segment_data = mapper.build_fbdi_row(
                transaction_transfer=transfer,
                base_row=journal_entry
            )
            
            sample_data.append(segment_data)

    # Create offsetting journal entry (TO side - credit entry)
    # Note: This assumes a control account/entity for the offsetting entry
    # In production, you'd need to determine the TO segments from business logic
    if total_debit > 0 and transfers:
         # Get first transfer to determine segment configuration
        first_transfer = transfers[0]
        
        # Base offsetting entry
        offsetting_entry = {
            "Status Code": "NEW",
            "Ledger ID": "300000205309206",
            "Effective Date of Transaction": "2025-09-17",
            "Journal Source": "Allocations",
            "Journal Category": "Adjustment",
            "Currency Code": "AED",
            "Journal Entry Creation Date": "2025-09-17",
            "Actual Flag": "E",
            "Entered Debit Amount": total_debit if type == "reject" else "",
            "Entered Credit Amount": total_debit if type == "submit" else "",
            "REFERENCE1 (Batch Name)": batch_name,
            "REFERENCE2 (Batch Description)": batch_description,
            "REFERENCE4 (Journal Entry Name)": journal_name,
            "REFERENCE5 (Journal Entry Description)": journal_description,
            "REFERENCE10 (Journal Entry Line Description)": "Offsetting credit line for balance transfer",
            "Encumbrance Type ID": "100000243328511",
            "Interface Group Identifier": group_id,
        }
        
        # Add dynamic segments for offsetting entry
        # Use the first transfer's TO segments (or build custom control account logic)
        offsetting_with_segments = mapper.build_fbdi_row(
            transaction_transfer=first_transfer,
            base_row=offsetting_entry,
            include_from_to='to'  # Use TO segments for offsetting entry
        )
        
        sample_data.append(offsetting_with_segments)

    return sample_data


def create_journal_entry_with_segments(
    segment_values: Dict[int, str],
    ledger_id: str = "300000205309206",
    effective_date: str = None,
    debit_amount: float = 0,
    credit_amount: float = 0,
    batch_name: str = None,
    journal_name: str = None,
    line_description: str = "",
    group_id: int = 0,
    **additional_fields
) -> Dict[str, Any]:
    """
    Create a single journal entry with custom segment values.

    Args:
        segment_values: Dict mapping segment_type_id to segment code
                       Example: {1: 'E001', 2: 'A100', 3: 'P001'}
        ledger_id: Oracle ledger ID
        effective_date: Transaction date (YYYY-MM-DD format)
        debit_amount: Debit amount (0 if credit entry)
        credit_amount: Credit amount (0 if debit entry)
        batch_name: Batch identifier
        journal_name: Journal entry identifier
        line_description: Description for this journal line
        group_id: Interface group identifier
        **additional_fields: Any other FBDI columns (e.g., Currency Code, Actual Flag)

    Returns:
        Dictionary with all FBDI columns including dynamic SEGMENT1-SEGMENT30

    Example:
        >>> entry = create_journal_entry_with_segments(
        ...     segment_values={1: 'E001', 2: 'A100', 3: 'P001'},
        ...     debit_amount=5000.00,
        ...     credit_amount=0,
        ...     line_description="Budget transfer"
        ... )
    """
    mapper = OracleSegmentMapper()
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    
    # Generate defaults if not provided
    if effective_date is None:
        effective_date = time.strftime("%Y-%m-%d")
    if batch_name is None:
        batch_name = f"BATCH_{timestamp}"
    if journal_name is None:
        journal_name = f"JOURNAL_{timestamp}"
    
    # Base journal entry structure
    base_entry = {
        "Status Code": "NEW",
        "Ledger ID": ledger_id,
        "Effective Date of Transaction": effective_date,
        "Journal Source": "Allocations",
        "Journal Category": "Adjustment",
        "Currency Code": "AED",
        "Journal Entry Creation Date": time.strftime("%Y-%m-%d"),
        "Actual Flag": "E",
        "Entered Debit Amount": debit_amount if debit_amount > 0 else "",
        "Entered Credit Amount": credit_amount if credit_amount > 0 else "",
        "REFERENCE1 (Batch Name)": batch_name,
        "REFERENCE2 (Batch Description)": f"Batch created {time.strftime('%Y-%m-%d %H:%M:%S')}",
        "REFERENCE4 (Journal Entry Name)": journal_name,
        "REFERENCE5 (Journal Entry Description)": f"Journal entry {time.strftime('%Y-%m-%d %H:%M:%S')}",
        "REFERENCE10 (Journal Entry Line Description)": line_description,
        "Encumbrance Type ID": "100000243328511",
        "Interface Group Identifier": group_id,
    }
    
    # Merge with additional fields
    base_entry.update(additional_fields)
    
    # Build segment columns dynamically
    # Add SEGMENT1 through SEGMENT30 based on segment_values
    for segment_type_id, segment_code in segment_values.items():
        oracle_field_name = mapper.get_oracle_field_name(segment_type_id)
        if oracle_field_name:
            base_entry[oracle_field_name] = segment_code
    
    # Fill remaining segments with empty/default values
    configured_segments = mapper.get_active_oracle_fields()
    all_segments = mapper.build_all_oracle_fields(num_segments=30)
    
    for segment_field in all_segments:
        if segment_field not in base_entry:
            # Use default empty value for unconfigured segments
            base_entry[segment_field] = "00000" if segment_field not in configured_segments else ""
    
    return base_entry


def create_balanced_journal_pair(
    debit_segments: Dict[int, str],
    credit_segments: Dict[int, str],
    amount: float,
    ledger_id: str = "300000205309206",
    effective_date: str = None,
    batch_name: str = None,
    journal_name: str = None,
    debit_description: str = "Debit entry",
    credit_description: str = "Credit entry",
    group_id: int = 0,
) -> List[Dict[str, Any]]:
    """
    Create a balanced pair of journal entries (debit + credit).

    Args:
        debit_segments: Segment values for debit entry
        credit_segments: Segment values for credit entry
        amount: Transaction amount (positive value)
        ledger_id: Oracle ledger ID
        effective_date: Transaction date
        batch_name: Batch identifier (auto-generated if None)
        journal_name: Journal identifier (auto-generated if None)
        debit_description: Description for debit line
        credit_description: Description for credit line
        group_id: Interface group identifier

    Returns:
        List of 2 journal entries (debit and credit)

    Example:
        >>> entries = create_balanced_journal_pair(
        ...     debit_segments={1: 'E001', 2: 'A100'},
        ...     credit_segments={1: 'E002', 2: 'A200'},
        ...     amount=10000.00
        ... )
    """
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    
    if batch_name is None:
        batch_name = f"BATCH_{timestamp}"
    if journal_name is None:
        journal_name = f"JOURNAL_{timestamp}"
    
    # Create debit entry
    debit_entry = create_journal_entry_with_segments(
        segment_values=debit_segments,
        ledger_id=ledger_id,
        effective_date=effective_date,
        debit_amount=amount,
        credit_amount=0,
        batch_name=batch_name,
        journal_name=journal_name,
        line_description=debit_description,
        group_id=group_id,
    )
    
    # Create credit entry
    credit_entry = create_journal_entry_with_segments(
        segment_values=credit_segments,
        ledger_id=ledger_id,
        effective_date=effective_date,
        debit_amount=0,
        credit_amount=amount,
        batch_name=batch_name,
        journal_name=journal_name,
        line_description=credit_description,
        group_id=group_id,
    )
    
    return [debit_entry, credit_entry]


if __name__ == "__main__":
    # Example usage with dynamic segments
    template_path = "JournalImportTemplate.xlsm"

    # Example 1: Create balanced journal pair with custom segments
    print("Example 1: Creating balanced journal pair with dynamic segments")
    journal_pair = create_balanced_journal_pair(
        debit_segments={1: 'E001', 2: 'A100', 3: 'P001'},  # Entity, Account, Project
        credit_segments={1: 'E002', 2: 'A200', 3: 'P002'},
        amount=15000.00,
        debit_description="Budget transfer FROM E001",
        credit_description="Budget transfer TO E002"
    )
    print(f"Created {len(journal_pair)} balanced entries")
    
    # Example 2: Create single entry with custom segments
    print("\nExample 2: Creating single journal entry")
    single_entry = create_journal_entry_with_segments(
        segment_values={
            1: 'E001',  # Entity
            2: 'A100',  # Account
            3: 'P001',  # Project
            4: 'M001',  # Department (if configured)
        },
        debit_amount=5000.00,
        credit_amount=0,
        line_description="Budget allocation"
    )
    print(f"Created entry with {len([k for k in single_entry.keys() if k.startswith('Segment')])} segment columns")

    # Combine for template filling
    sample_data = journal_pair + [single_entry]

    try:
        # Create journal from scratch with sample data
        result = create_journal_from_scratch(
            template_path=template_path,
            journal_data=sample_data,
            output_name=r"test_upload_fbdi\SampleJournal",
            auto_zip=True,
        )

        print(f"\nCompleted! Final file: {result}")
        print(f"Generated {len(sample_data)} journal entries with dynamic segments")

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
