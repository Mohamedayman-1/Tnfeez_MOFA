import pandas as pd
import openpyxl
from openpyxl import load_workbook
import shutil
from pathlib import Path
import time
from typing import Dict, List, Any
from test_upload_fbdi.zip_fbdi import excel_to_csv_and_zip
from test_upload_fbdi.upload_budget_fbdi import upload_budget_from_zip, upload_budget_fbdi_to_oracle
from test_upload_fbdi.budget_import_flow import upload_to_ucm , submit_interface_loader, submit_budget_import , run_budget_import_flow
import sys

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from account_and_entitys.oracle import OracleSegmentMapper

def create_clean_budget_template(template_path: str, output_path: str = None) -> str:
    """
    Create a clean copy of the BudgetImportTemplate with empty XCC_BUDGET_INTERFACE data rows.
    
    Args:
        template_path: Path to the original BudgetImportTemplate.xlsm
        output_path: Path for the clean template (optional, auto-generates if not provided)
    
    Returns:
        Path to the clean template file
    """
    template_file = Path(template_path)
    
    if not template_file.exists():
        raise FileNotFoundError(f"Template file not found: {template_path}")
    
    # Generate output path if not provided
    if output_path is None:
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        output_path = template_file.parent / f"BudgetImportTemplate_Clean_{timestamp}.xlsm"
    else:
        output_path = Path(output_path)
    
    # Create a copy of the template
    shutil.copy2(template_file, output_path)
    print(f"Created clean template copy: {output_path}")
    
    # Open the copy and clear XCC_BUDGET_INTERFACE data
    wb = load_workbook(output_path)
    
    if "XCC_BUDGET_INTERFACE" in wb.sheetnames:
        budget_sheet = wb["XCC_BUDGET_INTERFACE"]
        
        # Keep rows 1-4 (instructions and headers), delete data rows from row 5 onwards
        if budget_sheet.max_row > 4:
            budget_sheet.delete_rows(5, budget_sheet.max_row - 4)
            print(f"Cleared data rows from XCC_BUDGET_INTERFACE sheet, keeping header structure")
        
        # Save the changes
        wb.save(output_path)
        wb.close()
    else:
        wb.close()
        raise ValueError("XCC_BUDGET_INTERFACE sheet not found in template")
    
    return str(output_path)

def fill_budget_template_with_data(
    template_path: str, 
    budget_data: List[Dict[str, Any]], 
    output_path: str = None,
    auto_zip: bool = False
) -> str:
    """
    Fill a clean budget template with data and optionally create ZIP.
    
    Args:
        template_path: Path to the clean BudgetImportTemplate.xlsm
        budget_data: List of dictionaries containing budget entry data
        output_path: Path for the filled template (optional)
        auto_zip: Whether to automatically create ZIP file
    
    Returns:
        Path to the filled template file (or ZIP file if auto_zip=True)
    """
    template_file = Path(template_path)
    
    if not template_file.exists():
        raise FileNotFoundError(f"Template file not found: {template_path}")
    
    if not budget_data:
        raise ValueError("No budget data provided")
    
    # Generate output path if not provided
    if output_path is None:
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        output_path = template_file.parent / f"XccBudgetInterface_Filled_{timestamp}.xlsm"
    else:
        output_path = Path(output_path)
    
    # Create a copy for filling
    shutil.copy2(template_file, output_path)
    
    # Open the workbook and get the XCC_BUDGET_INTERFACE sheet
    wb = load_workbook(output_path)
    budget_sheet = wb["XCC_BUDGET_INTERFACE"]
    
    # Get the headers from row 4
    headers = []
    for col_num in range(1, budget_sheet.max_column + 1):
        header_cell = budget_sheet.cell(row=4, column=col_num)
        if header_cell.value is not None and str(header_cell.value).strip():
            clean_header = str(header_cell.value).lstrip('*').strip()
            headers.append(clean_header)
        else:
            break
    
    print(f"Found {len(headers)} headers in template")
    print(f"Filling template with {len(budget_data)} budget entries")
    
    # Fill data starting from row 5
    for row_idx, entry in enumerate(budget_data, start=5):
        for col_idx, header in enumerate(headers, start=1):
            # Get value from entry data
            value = entry.get(header, "")
            
            # Set the cell value
            cell = budget_sheet.cell(row=row_idx, column=col_idx)
            
            # Handle different data types appropriately
            if value is not None:
                cell.value = value
            else:
                cell.value = ""  # Empty for None
    
    # Save the filled template
    wb.save(output_path)
    wb.close()
    
    print(f"Template filled with data: {output_path}")
    
    # Optionally create ZIP file
    if auto_zip:
        zip_path = str(output_path).replace('.xlsm', '.zip')
        try:
            zip_result = excel_to_csv_and_zip(str(output_path), zip_path)
            print(f"ZIP file created: {zip_result}")
            return zip_result
        except Exception as e:
            print(f"Warning: Failed to create ZIP file: {e}")
            return str(output_path)
    
    return str(output_path)

def create_budget_from_scratch(
    template_path: str,
    budget_data: List[Dict[str, Any]],
    output_name: str = None,
    auto_zip: bool = True
) -> str:
    """
    Complete workflow: Create clean template, fill with data, and optionally ZIP.
    
    Args:
        template_path: Path to the original BudgetImportTemplate.xlsm
        budget_data: List of dictionaries containing budget entry data
        output_name: Base name for output files (optional) - will use XccBudgetInterface prefix if None
        auto_zip: Whether to create ZIP file automatically
    
    Returns:
        Path to the final file (Excel or ZIP)
    """
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    
    if output_name is None:
        # Use Oracle-required naming convention: XccBudgetInterface as prefix
        output_name = f"XccBudgetInterface_{timestamp}"
    
    try:
        # Step 1: Create clean template
        clean_template = create_clean_budget_template(template_path)
        
        # Step 2: Fill with data
        filled_template_path = Path(template_path).parent / f"{output_name}.xlsm"
        result_path = fill_budget_template_with_data(
            clean_template, 
            budget_data, 
            str(filled_template_path),
            auto_zip=auto_zip
        )
        
        # Clean up temporary clean template
        Path(clean_template).unlink(missing_ok=True)
        
        return result_path
    
    except Exception as e:
        print(f"Error in budget creation workflow: {e}")
        raise

def create_and_upload_budget(
    template_path: str,
    budget_data: List[Dict[str, Any]],
    output_name: str = None,
    upload_to_oracle: bool = True,
    group_id: str = None
) -> Dict[str, Any]:
    """
    Complete workflow: Create budget template, fill with data, ZIP, and upload to Oracle Fusion.
    
    Args:
        template_path: Path to the original BudgetImportTemplate.xlsm
        budget_data: List of dictionaries containing budget entry data
        output_name: Base name for output files (optional)
        upload_to_oracle: Whether to upload to Oracle Fusion via FBDI
        group_id: Optional group ID for Oracle upload
    
    Returns:
        Dictionary with workflow results including file paths and upload status
    """
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    
    if output_name is None:
        output_name = f"XccBudgetInterface_{timestamp}"
    
    try:
        # Step 1: Create budget ZIP file
        result_path = create_budget_from_scratch(
            template_path=template_path,
            budget_data=budget_data,
            output_name=output_name,
            auto_zip=True
        )
        
        workflow_result = {
            "success": True,
            "zip_file": result_path,
            "budget_entries": len(budget_data),
            "timestamp": timestamp
        }
        
        # Step 2: Upload to Oracle Fusion if requested
        if upload_to_oracle:
            print(f"Uploading budget to Oracle Fusion...")
            upload_result = run_budget_import_flow(result_path, transaction_id=701)
            
            workflow_result.update({
                "upload_success": upload_result["success"],
                "upload_error": upload_result.get("error"),
                "request_id": upload_result.get("request_id"),
                "group_id": upload_result.get("group_id"),
                "upload_message": upload_result.get("message")
            })
            
            if upload_result["success"]:
                print(f"✅ Budget uploaded successfully!")
                print(f"Request ID: {upload_result.get('request_id', 'N/A')}")
                print(f"Group ID: {upload_result.get('group_id', 'N/A')}")
            else:
                print(f"❌ Upload failed: {upload_result['error']}")
        else:
            workflow_result.update({
                "upload_success": None,
                "upload_message": "Upload skipped (upload_to_oracle=False)"
            })
        
        return workflow_result
        
    except Exception as e:
        return {
            "success": False,
            "error": f"Budget workflow failed: {str(e)}",
            "timestamp": timestamp
        }

# Example usage function
def create_sample_budget_data(transfers, transactions_id) -> List[Dict[str, Any]]:
    """
    Create budget entry data with dynamic segment support.
    
    Args:
        transfers: List of XX_TransactionTransfer objects with segment relationships
        transactions_id: Transaction identifier for budget entry naming
    
    Returns:
        List of budget entries with dynamic SEGMENT1-SEGMENT30 columns
    
    Notes:
        - Uses OracleSegmentMapper to dynamically build segment columns
        - Creates separate entries for FROM (negative) and TO (positive) amounts
        - Supports 2-30 segments based on transaction configuration
    """
    # Initialize mapper
    mapper = OracleSegmentMapper()
    
    # Generate batch identifiers
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    batch_name = "MIC_HQ_MONTHLY"
    budget_name = f"MIC_HQ_MONTHLY_{transactions_id}"
    
    sample_data = []
    line_number = 1
    
    for transfer in transfers:
        # Process FROM side (negative amount - debit)
        from_amount = getattr(transfer, 'from_center', 0) or 0
        if from_amount > 0:
            # Base budget entry
            budget_entry = {
                "Source Budget Type": "HYPERION",
                "Source Budget Name": batch_name,
                "Budget Entry Name": budget_name,
                "Line Number": line_number,
                "Amount": -1 * from_amount,  # Negative for FROM side
                "Currency Code": "AED",
                "Period Name": "Sep-25",
            }
            
            # Add dynamic segment columns using mapper
            # This builds SEGMENT1, SEGMENT2, ... SEGMENT30 based on FROM segments
            budget_entry_with_segments = mapper.build_fbdi_row(
                transaction_transfer=transfer,
                base_row=budget_entry,
                include_from_to='from'  # Use FROM segments
            )
            
            sample_data.append(budget_entry_with_segments)
            line_number += 1
        
        # Process TO side (positive amount - credit)
        to_amount = getattr(transfer, 'to_center', 0) or 0
        if to_amount > 0:
            # Base budget entry
            budget_entry = {
                "Source Budget Type": "HYPERION",
                "Source Budget Name": batch_name,
                "Budget Entry Name": budget_name,
                "Line Number": line_number,
                "Amount": to_amount,  # Positive for TO side
                "Currency Code": "AED",
                "Period Name": "Sep-25",
            }
            
            # Add dynamic segment columns using mapper
            # This builds SEGMENT1, SEGMENT2, ... SEGMENT30 based on TO segments
            budget_entry_with_segments = mapper.build_fbdi_row(
                transaction_transfer=transfer,
                base_row=budget_entry,
                include_from_to='to'  # Use TO segments
            )
            
            sample_data.append(budget_entry_with_segments)
            line_number += 1
    
    return sample_data


def create_budget_entry_with_segments(
    segment_values: Dict[int, str],
    amount: float,
    budget_name: str,
    line_number: int = 1,
    source_budget_type: str = "HYPERION",
    source_budget_name: str = "MIC_HQ_MONTHLY",
    period_name: str = "Sep-25",
    currency_code: str = "AED",
) -> Dict[str, Any]:
    """
    Create a single budget entry with custom segment values.

    Args:
        segment_values: Dict mapping segment_type_id to segment code
                       Example: {1: 'E001', 2: 'A100', 3: 'P001'}
        amount: Budget amount (positive for increase, negative for decrease)
        budget_name: Budget entry name identifier
        line_number: Line number within budget entry
        source_budget_type: Budget type (e.g., "HYPERION", "ORACLE")
        source_budget_name: Source budget identifier
        period_name: Period name (e.g., "Sep-25", "Oct-25")
        currency_code: Currency code (default: AED)

    Returns:
        Dictionary with all FBDI columns including dynamic SEGMENT1-SEGMENT30

    Example:
        >>> entry = create_budget_entry_with_segments(
        ...     segment_values={1: 'E001', 2: 'A100', 3: 'P001'},
        ...     amount=10000.00,
        ...     budget_name="Q3_BUDGET_TRANSFER",
        ...     line_number=1
        ... )
    """
    mapper = OracleSegmentMapper()
    
    # Base budget entry structure
    base_entry = {
        "Source Budget Type": source_budget_type,
        "Source Budget Name": source_budget_name,
        "Budget Entry Name": budget_name,
        "Line Number": line_number,
        "Amount": amount,
        "Currency Code": currency_code,
        "Period Name": period_name,
    }
    
    # Build segment columns dynamically
    for segment_type_id, segment_code in segment_values.items():
        oracle_field_name = mapper.get_oracle_field_name(segment_type_id)
        if oracle_field_name:
            base_entry[oracle_field_name] = str(segment_code)
    
    # Fill remaining segments with empty values for unconfigured segments
    configured_segments = mapper.get_active_oracle_fields()
    all_segments = mapper.build_all_oracle_fields(num_segments=30)
    
    for segment_field in all_segments:
        if segment_field not in base_entry:
            # Leave unconfigured segments empty (Oracle will use defaults)
            base_entry[segment_field] = ""
    
    return base_entry


def create_budget_transfer_pair(
    from_segments: Dict[int, str],
    to_segments: Dict[int, str],
    amount: float,
    budget_name: str,
    period_name: str = "Sep-25",
    starting_line_number: int = 1,
) -> List[Dict[str, Any]]:
    """
    Create a balanced pair of budget entries (FROM negative + TO positive).

    Args:
        from_segments: Segment values for FROM entry (debit/decrease)
        to_segments: Segment values for TO entry (credit/increase)
        amount: Transfer amount (positive value)
        budget_name: Budget entry identifier
        period_name: Period name (e.g., "Sep-25")
        starting_line_number: Starting line number for entries

    Returns:
        List of 2 budget entries (FROM negative, TO positive)

    Example:
        >>> entries = create_budget_transfer_pair(
        ...     from_segments={1: 'E001', 2: 'A100'},
        ...     to_segments={1: 'E002', 2: 'A200'},
        ...     amount=5000.00,
        ...     budget_name="MONTHLY_REALLOCATION"
        ... )
    """
    # Create FROM entry (negative amount)
    from_entry = create_budget_entry_with_segments(
        segment_values=from_segments,
        amount=-1 * amount,  # Negative for decrease
        budget_name=budget_name,
        line_number=starting_line_number,
        period_name=period_name,
    )
    
    # Create TO entry (positive amount)
    to_entry = create_budget_entry_with_segments(
        segment_values=to_segments,
        amount=amount,  # Positive for increase
        budget_name=budget_name,
        line_number=starting_line_number + 1,
        period_name=period_name,
    )
    
    return [from_entry, to_entry]


if __name__ == "__main__":
    # Example usage with dynamic segments
    template_path = "BudgetImportTemplate.xlsm"
    
    print("=== Budget Template Manager - Dynamic Segment Examples ===\n")
    
    # Example 1: Create budget transfer pair with custom segments
    print("Example 1: Creating budget transfer pair with dynamic segments")
    budget_pair = create_budget_transfer_pair(
        from_segments={1: 'E001', 2: 'A100', 3: 'P001'},  # FROM: Entity E001, Account A100, Project P001
        to_segments={1: 'E002', 2: 'A200', 3: 'P002'},    # TO: Entity E002, Account A200, Project P002
        amount=25000.00,
        budget_name="MONTHLY_BUDGET_TRANSFER",
        period_name="Sep-25"
    )
    print(f"Created {len(budget_pair)} balanced budget entries")
    
    # Example 2: Create single budget entry
    print("\nExample 2: Creating single budget entry with custom segments")
    single_entry = create_budget_entry_with_segments(
        segment_values={
            1: 'E001',  # Entity
            2: 'A100',  # Account
            3: 'P001',  # Project
            4: 'D001',  # Department (if configured)
        },
        amount=10000.00,
        budget_name="Q3_BUDGET_ADJUSTMENT",
        line_number=1,
        period_name="Sep-25"
    )
    print(f"Created entry with {len([k for k in single_entry.keys() if k.startswith('Segment')])} segment columns")
    
    # Combine for template filling
    sample_data = budget_pair + [single_entry]
    
    try:
        print("\n=== Budget Creation and Upload Workflow ===")
        print(f"Processing {len(sample_data)} budget entries with dynamic segments...")
        
        # Option 1: Create budget and upload to Oracle in one step
        result = create_and_upload_budget(
            template_path=template_path,
            budget_data=sample_data,
            output_name="XccBudgetInterface_DynamicSegments",
            upload_to_oracle=False,  # Set to True to upload to Oracle
            group_id=f"BUDGET_TEST_{time.strftime('%Y%m%d_%H%M%S')}"
        )
        
        print(f"\n=== Workflow Results ===")
        print(f"Success: {result['success']}")
        if result['success']:
            print(f"ZIP File: {result['zip_file']}")
            print(f"Budget Entries: {result['budget_entries']}")
            if result.get('upload_success'):
                print(f"Upload Status: ✅ Success")
                print(f"Request ID: {result.get('request_id', 'N/A')}")
                print(f"Group ID: {result.get('group_id', 'N/A')}")
            elif result.get('upload_success') is False:
                print(f"Upload Status: ❌ Failed - {result.get('upload_error', 'Unknown error')}")
            else:
                print(f"Upload Status: ⏭️ Skipped (set upload_to_oracle=True to enable)")
        else:
            print(f"Error: {result.get('error', 'Unknown error')}")
        
        print(f"\nGenerated {len(sample_data)} budget entries with dynamic segment support")
        print("✅ Budget entries support 2-30 segments based on configuration")
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()